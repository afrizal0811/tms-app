from openpyxl.styles import Alignment, PatternFill, Font
import pandas as pd
import numpy as np # Import numpy for sorting
import re
import requests
import traceback
import math
import datetime
from datetime import timedelta
from openpyxl.comments import Comment
from utils.function import (
    get_save_path,
    load_config,
    load_constants,
    load_master_data,
    load_secret,
    open_file_externally
)
from utils.gui import create_date_picker_window
from utils.function import show_error_message, show_info_message
from utils.messages import ERROR_MESSAGES, INFO_MESSAGES
from utils.api_handler import handle_requests_error
import json
def haversine_distance(coord1_str, coord2_str):
    """
    Menghitung jarak Haversine antara dua koordinat (lat, long) dalam meter.
    Koordinat harus dalam format string "lat,long".
    Mengembalikan 0 jika salah satu koordinat tidak valid atau kosong.
    """
    if not coord1_str or not coord2_str:
        return 0

    try:
        lat1, lon1 = map(float, coord1_str.split(','))
        lat2, lon2 = map(float, coord2_str.split(','))
    except ValueError:
        return 0

    R = 6371000
    lat1_rad, lon1_rad, lat2_rad, lon2_rad = map(math.radians, [lat1, lon1, lat2, lon2])
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    distance = R * c
    return round(distance)

def process_task_data_code1(task, master_map, real_sequence_map):
    """
    Memproses satu data 'task' (Logic from Code 1).
    Returns None if essential assignee info is missing.
    """
    
    # --- PERBAIKAN: Gunakan assignedTo SEBAGAI FALLBACK ---
    vehicle_assignee_email = (task.get('assignedVehicle') or {}).get('assignee')
    assigned_to_data = task.get('assignedTo') or {}
    assigned_to_email = assigned_to_data.get('email')
    assigned_to_name = assigned_to_data.get('name') # Nama driver dari assignedTo

    # Tentukan email utama (prioritas vehicle, fallback ke assignedTo)
    final_assignee_email = vehicle_assignee_email or assigned_to_email
    
    # Filter Awal: Jika tidak ada email assignee SAMA SEKALI, skip task ini
    if not final_assignee_email:
        return None

    master_record = master_map.get(final_assignee_email, {})
    
    # Tentukan nama driver (prioritas assignedTo.name, fallback ke master, fallback ke email)
    driver_name = assigned_to_name or master_record.get('Driver', final_assignee_email)
    
    # Tentukan plat (prioritas assignedVehicle.name, fallback ke master)
    license_plat = (task.get('assignedVehicle') or {}).get('name')
    if not license_plat or license_plat == 'N/A':
        license_plat = master_record.get('Plat', 'N/A')

    customer_name = task.get('customerName', '')

    t_arrival_utc = pd.to_datetime(task.get('klikJikaSudahSampai'), errors='coerce')
    # Gunakan 'doneTime' sebagai waktu departure untuk konsistensi
    t_departure_utc = pd.to_datetime(task.get('doneTime'), errors='coerce')
    t_arrival_local = t_arrival_utc.tz_convert('Asia/Jakarta') if pd.notna(t_arrival_utc) else pd.NaT
    t_departure_local = t_departure_utc.tz_convert('Asia/Jakarta') if pd.notna(t_departure_utc) else pd.NaT

    actual_visit_time = pd.NA
    if pd.notna(t_arrival_local) and pd.notna(t_departure_local):
        delta_minutes = (t_departure_local - t_arrival_local).total_seconds() / 60
        actual_visit_time = int(round(delta_minutes))

    et_sequence = task.get('routePlannedOrder') # Bisa None
    et_sequence_val = et_sequence if et_sequence is not None else 0 # Default ke 0 jika None

    real_sequence = real_sequence_map.get(task['_id'], 0) # Real sequence dari map (Code 1 logic)

    # --- PERBAIKAN (dari chat sebelumnya): Gabungkan SEMUA sumber status ---
    status_delivery_1 = task.get('statusDelivery', [])
    status_delivery_2 = task.get('statusGr', [])
    raw_labels = task.get('label')

    # Pastikan status_delivery_1 & 2 adalah list (API bisa saja 'None')
    if not isinstance(status_delivery_1, list): status_delivery_1 = []
    if not isinstance(status_delivery_2, list): status_delivery_2 = []
    
    # Gabungkan list awal
    combined_list = status_delivery_1 + status_delivery_2
    
    # Tambahkan 'label' ke combined_list
    if isinstance(raw_labels, str):
        combined_list.append(raw_labels) # Tambahkan string
    elif isinstance(raw_labels, list):
        combined_list.extend(raw_labels) # Tambahkan list
        
    # Buat list unik (dan hapus string kosong/None)
    final_status_list = list(dict.fromkeys(s for s in combined_list if s)) 
    
    # Tetapkan hasil ke *kedua* variabel untuk konsistensi
    labels_list = final_status_list
    combined_status_delivery = final_status_list
    return {
        'task_id': task['_id'],
        'flow': task.get('flow', ''),
        'license_plat': license_plat, # Variabel baru
        'driver_name': driver_name, # Variabel baru
        'assignee_email': final_assignee_email, # Variabel baru
        'customer_name': customer_name,
        'status_delivery': ', '.join(combined_status_delivery), # Ini tetap dipakai Pending SO
        'status_delivery_list': combined_status_delivery, # Ini tetap dipakai Pending SO
        'open_time': task.get('openTime', ''),
        'close_time': task.get('closeTime', ''),
        'eta': (task.get('eta') or '')[:5],
        'etd': (task.get('etd') or '')[:5],
        'actual_arrival': t_arrival_local.strftime('%H:%M') if pd.notna(t_arrival_local) else '',
        'actual_departure': t_departure_local.strftime('%H:%M') if pd.notna(t_departure_local) else '',
        'visit_time': task.get('visitTime', ''),
        'actual_visit_time': actual_visit_time,
        'et_sequence': et_sequence_val,
        'real_sequence': real_sequence,
        'is_same_sequence': "SAMA" if et_sequence is not None and et_sequence_val == real_sequence else "TIDAK SAMA",
        'labels': labels_list,
        'alasan': task.get('alasan', ''),
        # Tambahan untuk perhitungan Real Sequence & Actual Visit Time di RO vs Real
        '_arrival_dt_local': t_arrival_local,
        '_departure_dt_local': t_departure_local,
        # Tambahkan _arrival_utc untuk sorting RO vs Real di Code 2 logic
        '_arrival_utc': t_arrival_utc
    }
# =============================================================================

def format_excel_sheet(writer, df, sheet_name, centered_cols, colored_cols=None):
    """Menulis DataFrame ke sheet dan menerapkan format."""
    df.to_excel(writer, index=False, sheet_name=sheet_name)
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    center_align = Alignment(horizontal='center', vertical='center')

    for idx, col_name in enumerate(df.columns):
        col_letter = chr(65 + idx)
        try:
            # Hitung panjang data dan header
            data_lengths = df[col_name].astype(str).map(len)
            max_len_data = data_lengths.max() if not data_lengths.empty else 0
            header_len = len(str(col_name))
            max_len = max(max_len_data, header_len) + 2
            worksheet.column_dimensions[col_letter].width = max_len
        except (ValueError, TypeError, AttributeError): # Tangani error
             worksheet.column_dimensions[col_letter].width = len(str(col_name)) + 2 # Fallback

        if col_name in centered_cols:
            for cell in worksheet[col_letter][1:]: # Mulai dari baris ke-2
                if cell.value is not None and cell.value != '':
                    cell.alignment = center_align

        # --- PERBAIKAN DI BLOK INI ---
        if colored_cols and col_name in colored_cols:
            fill = PatternFill(start_color=colored_cols[col_name], end_color=colored_cols[col_name], fill_type="solid")
            for cell in worksheet[col_letter]:
                 # Kondisi 'if cell.value' dihapus agar seluruh kolom diwarnai
                 cell.fill = fill

    header_align = Alignment(horizontal='center', vertical='center')
    for cell in worksheet[1]: # Hanya header
        cell.alignment = header_align

def fetch_results_data(base_url, headers, date_str, hub_id):
    """
    Mengambil data dari endpoint /results untuk mendapatkan ETA/ETD hub.
    """
    results_url = f"{base_url}/results"
    params = {
        'dateFrom': date_str,
        'dateTo': date_str,
        'hubId': hub_id,
        'limit': 500  # Asumsi 500 data cukup untuk 1 hari
    }

    try:
        response = requests.get(results_url, headers=headers, params=params, timeout=60)
        response.raise_for_status()  # Cek error HTTP
        return response.json()
    except requests.exceptions.RequestException as e:
        show_error_message("Gagal API /results", f"Gagal mengambil data dari {results_url}: {e}\n\n{traceback.format_exc()}")
        return None

# GANTI FUNGSI LAMA DENGAN FUNGSI BARU INI
def parse_hub_times(results_data, master_map):
    """
    (Versi Bersih - v7)
    Mem-parsing JSON dari /results untuk membuat map:
    { "DriverName": {"first_etd": "HH:MM", "last_eta": "HH:MM"} }
    """
    hub_times_map = {}
    
    if not results_data or 'data' not in results_data or 'data' not in results_data['data']:
        return hub_times_map

    dispatch_list = results_data['data']['data']
    if not dispatch_list:
         return hub_times_map

    # Iterasi utama (per dispatch)
    for i, dispatch in enumerate(dispatch_list):
        
        if dispatch.get('dispatchStatus') != 'done':
            continue
        
        if 'result' not in dispatch or 'routing' not in dispatch['result']:
            continue
        
        # Iterasi vehicle ('routing')
        for j, vehicle in enumerate(dispatch['result']['routing']):
            
            # Filter 4: Cek 'assignee' (email)
            assignee_email = vehicle.get('assignee')
            if not assignee_email:
                continue

            # Mapping email ke Nama Driver
            master_record = master_map.get(assignee_email)
            driver_name = None
            
            if master_record is not None:
                driver_name = master_record.get('Driver') 
            
            if not driver_name:
                continue
            
            if driver_name in hub_times_map:
                continue
            
            trips = vehicle.get('trips', [])
            if not trips:
                continue

            first_hub_etd = None
            last_hub_eta = None

            # 1. Cari ETD dari Hub Pertama (Loop dari Awal)
            for visit in trips:
                if visit.get('isHub') is True:
                    etd = visit.get('etd')
                    if etd:
                        first_hub_etd = etd[:5] # Ambil HH:mm
                        break # Stop setelah menemukan yg pertama

            # 2. Cari ETA dari Hub Terakhir (Loop dari Akhir)
            for visit in reversed(trips):
                if visit.get('isHub') is True:
                    eta = visit.get('eta')
                    if eta:
                        last_hub_eta = eta[:5] # Ambil HH:mm
                        break # Stop setelah menemukan yg terakhir
            
            # 3. Simpan jika salah satu atau keduanya ditemukan
            if first_hub_etd or last_hub_eta:
                hub_times_map[driver_name] = {
                    'first_etd': first_hub_etd or '',
                    'last_eta': last_hub_eta or ''
                }

    return hub_times_map

def panggil_api_dan_simpan(dates, app_instance):
    """
    Fungsi utama untuk memanggil API, memproses data, dan menyimpan ke Excel.
    """
    selected_date = dates["ymd"]
    
    # --- PENGATURAN MENGGUNAKAN SHARED UTILS ---
    constants = load_constants()
    config = load_config()
    secrets = load_secret()
    master_data = load_master_data()

    master_df = master_data["df"]
    hub_ids = master_data["hub_ids"]

    if not constants: show_error_message("Gagal", ERROR_MESSAGES["CONSTANT_FILE_ERROR"]); return False
    if not config: show_error_message("Gagal", ERROR_MESSAGES["CONFIG_FILE_ERROR"]); return False
    if not secrets: show_error_message("Gagal", ERROR_MESSAGES["SECRET_FILE_ERROR"]); return False
    if master_data is None: show_error_message("Gagal", ERROR_MESSAGES["MASTER_DATA_MISSING"]); return False

    master_map = {row['Email']: row for _, row in master_df.iterrows()}

    API_TOKEN = secrets.get('token')
    LOKASI_FILTER = config.get('lokasi')
    HUB_ID = hub_ids.get(LOKASI_FILTER)
    location_id = constants.get('location_id', {})
    show_pending_gr = LOKASI_FILTER in ["plck", "pldm"]

    if not API_TOKEN: show_error_message("Error Token API", ERROR_MESSAGES["API_TOKEN_MISSING"]); return False
    if not LOKASI_FILTER or not HUB_ID: show_error_message("Konfigurasi Salah", ERROR_MESSAGES["HUB_ID_MISSING"]); return False

    base_url = constants.get('base_url')
    headers = {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}
    
    # --- LOGIKA PANGGIL API /RESULTS (H-1 / H-2) ---
    try:
        date_obj = datetime.datetime.strptime(selected_date, '%Y-%m-%d').date()
    except ValueError:
        show_error_message("Format Tanggal Salah", f"Format tanggal {selected_date} tidak valid.")
        return False
    
    day_of_week = date_obj.weekday()
    if day_of_week == 0: # 0 adalah Senin
        target_date_obj = date_obj - timedelta(days=2) # Mundur 2 hari (ke Sabtu)
    else:
        target_date_obj = date_obj - timedelta(days=1) # Mundur 1 hari
    
    results_date_str = target_date_obj.strftime('%Y-%m-%d')
    
    results_data = fetch_results_data(base_url, headers, results_date_str, HUB_ID)
    
    hub_times_map = parse_hub_times(results_data, master_map)

    # --- Panggilan API /tasks (Existing) ---
    api_url = f"{base_url}/tasks"
    params = {
        "status": "DONE", "hubId": HUB_ID,
        "timeFrom": f"{selected_date} 00:00:00", "timeTo": f"{selected_date} 23:59:59",
        "timeBy": "doneTime", "limit": 1000
    }

    try:
        response = requests.get(api_url, headers=headers, params=params, timeout=60)
        response.raise_for_status()
        tasks_data = response.json().get('tasks', {}).get('data')

        if not tasks_data:
            show_error_message("Data Tidak Ditemukan", ERROR_MESSAGES["DATA_NOT_FOUND"])
            return False

    except requests.exceptions.RequestException as e: handle_requests_error(e); return False
    except Exception as e: show_error_message("Error API", ERROR_MESSAGES["UNKNOWN_ERROR"].format(error_detail=f"{e}\n\n{traceback.format_exc()}")); return False

    # --- Data Processing (Tidak diubah) ---
    tasks_by_assignee_for_seq = {}
    for task in tasks_data:
        assignee_email_vehicle = (task.get('assignedVehicle') or {}).get('assignee')
        if assignee_email_vehicle and LOKASI_FILTER in assignee_email_vehicle:
            tasks_by_assignee_for_seq.setdefault(assignee_email_vehicle, []).append(task)

    real_sequence_map = {}
    for assignee, tasks in tasks_by_assignee_for_seq.items():
        sorted_tasks = sorted(tasks, key=lambda x: x.get('doneTime') or '9999-12-31T23:59:59Z')
        for i, task in enumerate(sorted_tasks):
            real_sequence_map[task['_id']] = i + 1

    summary_data_total_delivered_new = {}
    ro_vs_real_raw_data = {} 
    processed_tasks_list_pending = [] 
    pending_undelivered_labels = ["PENDING", "BATAL", "TERIMA SEBAGIAN", "PENDING GR"]
    if show_pending_gr: pending_undelivered_labels.append("PENDING GR")
    for task in tasks_data:
        assigned_to_data = task.get('assignedTo')
        driver_name_from_assigned_to = None
        driver_email_from_assigned_to = None
        if isinstance(assigned_to_data, dict):
            driver_name_from_assigned_to = assigned_to_data.get('name')
            driver_email_from_assigned_to = assigned_to_data.get('email')
        if driver_name_from_assigned_to:
            if driver_name_from_assigned_to not in summary_data_total_delivered_new:
                plat_td = "N/A_Plat"
                if driver_email_from_assigned_to and driver_email_from_assigned_to in master_map:
                    plat_td = master_map[driver_email_from_assigned_to].get('Plat', 'N/A_Plat')
                elif task.get('assignedVehicle') and isinstance(task['assignedVehicle'], dict):
                     plat_td = task['assignedVehicle'].get('name', 'N/A_Plat')
                summary_data_total_delivered_new[driver_name_from_assigned_to] = {
                    'License Plat': plat_td, 'Driver': driver_name_from_assigned_to,
                    'Total Outlet': 0, 'Total Delivered': 0
                }
            summary_data_total_delivered_new[driver_name_from_assigned_to]['Total Outlet'] += 1
            raw_labels_td = task.get('label')
            labels_td_list = []
            if isinstance(raw_labels_td, str): labels_td_list = [raw_labels_td]
            elif isinstance(raw_labels_td, list): labels_td_list = raw_labels_td
            failure_labels_td = ["PENDING", "BATAL", "TERIMA SEBAGIAN"]
            if not show_pending_gr: failure_labels_td.append("PENDING GR")
            is_pending_or_batal_td = any(label in failure_labels_td for label in labels_td_list)
            if not is_pending_or_batal_td:
                summary_data_total_delivered_new[driver_name_from_assigned_to]['Total Delivered'] += 1
            plat_ro = "N/A_Plat"
            if driver_email_from_assigned_to and driver_email_from_assigned_to in master_map:
                plat_ro = master_map[driver_email_from_assigned_to].get('Plat', 'N/A_Plat')
            elif task.get('assignedVehicle') and isinstance(task['assignedVehicle'], dict):
                 plat_ro = task['assignedVehicle'].get('name', 'N/A_Plat')
            customer_name_ro = task.get('customerName', '')
            raw_labels_ro = task.get('label') 
            if isinstance(raw_labels_ro, str): status_delivery_ro = raw_labels_ro
            elif isinstance(raw_labels_ro, list): status_delivery_ro = ', '.join(raw_labels_ro)
            else: status_delivery_ro = ''
            open_time_ro = task.get('openTime', '')
            close_time_ro = task.get('closeTime', '')
            flow_ro = task.get('flow', '') 
            arrival_key = 'page1DoneTime' if 'Pending GR' in flow_ro else 'klikJikaSudahSampai'
            arrival_utc_ro = pd.to_datetime(task.get(arrival_key), errors='coerce')
            departure_utc_ro = pd.to_datetime(task.get('doneTime'), errors='coerce')
            visit_time_api_ro = task.get('visitTime', '')
            ro_sequence_ro = task.get('routePlannedOrder') 
            eta_ro = (task.get('eta') or '')[:5]
            etd_ro = (task.get('etd') or '')[:5]
            task_details_for_ro = {
                '_task_id': task['_id'], 'Flow': flow_ro, 'Plat': plat_ro,
                'Driver': driver_name_from_assigned_to, 'Customer': customer_name_ro,
                'Status Delivery': status_delivery_ro, 'Open Time': open_time_ro,
                'Close Time': close_time_ro, 'ETA': eta_ro, 'ETD': etd_ro,
                'Visit Time': visit_time_api_ro,
                'RO Sequence': ro_sequence_ro if ro_sequence_ro is not None else '-',
                '_arrival_utc': arrival_utc_ro, '_departure_utc': departure_utc_ro
            }
            if driver_name_from_assigned_to not in ro_vs_real_raw_data:
                ro_vs_real_raw_data[driver_name_from_assigned_to] = []
            ro_vs_real_raw_data[driver_name_from_assigned_to].append(task_details_for_ro)
        processed_code1 = process_task_data_code1(task, master_map, real_sequence_map) 
        if processed_code1 and LOKASI_FILTER in processed_code1['assignee_email']:
            if ( any(label in pending_undelivered_labels for label in processed_code1['labels'])
                or any(status in pending_undelivered_labels for status in processed_code1.get('status_delivery_list', [])) ):
                 processed_tasks_list_pending.append(processed_code1)

    df_delivered = pd.DataFrame(list(summary_data_total_delivered_new.values()))
    if not df_delivered.empty:
        df_delivered['is_sewa'] = (df_delivered['License Plat'].astype(str).str.contains('SEWA', case=False, na=False) | df_delivered['Driver'].astype(str).str.contains('SEWA', case=False, na=False)).astype(int)
        conditions = [df_delivered['Driver'].astype(str).str.contains('DRY', case=False, na=False), df_delivered['Driver'].astype(str).str.contains('FRZ', case=False, na=False)]
        choices = [1, 2]
        df_delivered['sewa_category'] = np.select(conditions, choices, default=3)
        df_delivered = df_delivered.sort_values(by=['is_sewa', 'sewa_category', 'Driver'], ascending=[True, True, True]).reset_index(drop=True)
        df_delivered = df_delivered.drop(columns=['is_sewa', 'sewa_category'])

    ro_vs_real_final_list = []
    correct_sequence_map = {}
    for driver_name, tasks_list in ro_vs_real_raw_data.items():
        tasks_sorted_by_arrival = sorted(tasks_list, key=lambda x: x['_arrival_utc'] if pd.notna(x['_arrival_utc']) else pd.Timestamp.max.tz_localize('UTC'))
        arrival_rank_map = {}
        for i, task in enumerate(tasks_sorted_by_arrival):
            arrival_rank_map[task['_task_id']] = i + 1
        for task_detail in tasks_list:
            task_detail['_real_sequence_rank'] = arrival_rank_map.get(task_detail['_task_id'], 999) 
        sorted_tasks_for_display = sorted(tasks_list, key=lambda x: x['ETA'] if x['ETA'] else '99:99')
        for i, task_detail in enumerate(sorted_tasks_for_display):
            real_sequence = task_detail['_real_sequence_rank'] 
            arrival_local = task_detail['_arrival_utc'].tz_convert('Asia/Jakarta') if pd.notna(task_detail['_arrival_utc']) else pd.NaT
            departure_local = task_detail['_departure_utc'].tz_convert('Asia/Jakarta') if pd.notna(task_detail['_departure_utc']) else pd.NaT
            actual_visit_time_ro = pd.NA
            if pd.notna(arrival_local) and pd.notna(departure_local):
                 delta_minutes = (departure_local - arrival_local).total_seconds() / 60
                 actual_visit_time_ro = int(round(delta_minutes))
            ro_sequence_val = task_detail['RO Sequence']
            is_same = "SAMA" if ro_sequence_val != '-' and pd.to_numeric(ro_sequence_val, errors='coerce') == real_sequence else "TIDAK SAMA"
            task_id = task_detail.get('_task_id')
            if task_id:
                correct_sequence_map[task_id] = { 'ro': ro_sequence_val, 'real': real_sequence }
            ro_vs_real_final_list.append({
                'Flow': task_detail['Flow'], 'Plat': task_detail['Plat'],
                'Driver': task_detail['Driver'], 'Customer': task_detail['Customer'],
                'Status Delivery': task_detail['Status Delivery'], 'Open Time': task_detail['Open Time'],
                'Close Time': task_detail['Close Time'], 'ETA': task_detail['ETA'], 
                'Actual Arrival': arrival_local.strftime('%H:%M') if pd.notna(arrival_local) else '',
                'ETD': task_detail['ETD'], 
                'Actual Departure': departure_local.strftime('%H:%M') if pd.notna(departure_local) else '',
                'Visit Time': task_detail['Visit Time'], 'Actual Visit Time': actual_visit_time_ro,
                'RO Sequence': ro_sequence_val, 'Real Sequence': real_sequence, 
                'Is Same Sequence': is_same
            })

    df_ro_vs_real = pd.DataFrame(ro_vs_real_final_list)

    if not df_ro_vs_real.empty:
        df_ro_vs_real['ETA_Sort'] = df_ro_vs_real['ETA'].replace('', '99:99')
        df_ro_vs_real = df_ro_vs_real.sort_values(by=['Driver', 'ETA_Sort'], ascending=[True, True])
        df_ro_vs_real = df_ro_vs_real.drop(columns=['ETA_Sort'])
        
        final_ro_dfs = [] 
        ro_cols_order = ['Flow', 'Plat', 'Driver', 'Customer', 'Status Delivery', 'Open Time', 'Close Time', 'ETA', 'Actual Arrival', 'ETD', 'Actual Departure', 'Visit Time', 'Actual Visit Time', 'RO Sequence', 'Real Sequence', 'Is Same Sequence']
        
        df_ro_vs_real = df_ro_vs_real.reindex(columns=ro_cols_order)

        for driver_name, driver_group_df in df_ro_vs_real.groupby('Driver', sort=False):
            if not driver_name: 
                continue
            
            hub_data = hub_times_map.get(driver_name, {})
            hub_etd = hub_data.get('first_etd', '') 
            hub_eta = hub_data.get('last_eta', '')
            
            # 2. Buat baris HUB pertama (Hapus prefix "HUB:")
            hub_row_start = {col: '' for col in ro_cols_order}
            hub_row_start['Customer'] = 'HUB'
            # --- MODIFIKASI: Hapus prefix "HUB:" ---
            hub_row_start['ETD'] = hub_etd if hub_etd else '-' 
            
            # 3. Buat baris HUB terakhir (Hapus prefix "HUB:")
            hub_row_end = {col: '' for col in ro_cols_order}
            hub_row_end['Customer'] = 'HUB'
            # --- MODIFIKASI: Hapus prefix "HUB:" ---
            hub_row_end['ETA'] = hub_eta if hub_eta else '-'
            
            # 4. Buat DataFrame dari baris
            df_hub_start = pd.DataFrame([hub_row_start], columns=ro_cols_order)
            df_hub_end = pd.DataFrame([hub_row_end], columns=ro_cols_order)
            
            # 5. Buat baris Spacer
            spacer_row = pd.DataFrame([{col: '' for col in ro_cols_order}], columns=ro_cols_order)
            
            if final_ro_dfs:
                final_ro_dfs.append(spacer_row)
                
            final_ro_dfs.append(df_hub_start)
            final_ro_dfs.append(driver_group_df[ro_cols_order])
            final_ro_dfs.append(df_hub_end)

        if final_ro_dfs:
            df_ro_vs_real = pd.concat(final_ro_dfs, ignore_index=True)
        else:
            df_ro_vs_real = pd.DataFrame(columns=ro_cols_order)

    # --- Finalisasi Sheet 'Hasil Pending SO' (Tidak diubah) ---
    pending_so_data = [] 
    fill_values = None   
    df_pending = pd.DataFrame() 
    for processed in processed_tasks_list_pending: 
        labels_list = processed.get('labels', []) + processed.get('status_delivery_list', [])
        if not any(label in pending_undelivered_labels for label in labels_list): continue
        task_id = processed['task_id']
        correct_seqs = correct_sequence_map.get(task_id, {})
        ro_sequence = correct_seqs.get('ro', processed['et_sequence'])
        real_sequence = correct_seqs.get('real', processed['real_sequence'])
        match = re.search(r'(C0[0-9]+)', processed['customer_name'])
        reason = '' 
        if any(label in pending_undelivered_labels for label in labels_list):
            reason = processed['alasan']
        is_pending_gr = "PENDING GR" in labels_list
        is_pending = "PENDING" in labels_list
        is_batal = "BATAL" in labels_list
        is_sebagian = "TERIMA SEBAGIAN" in labels_list
        fill_color = None
        is_redirected_gr = (is_pending_gr and not show_pending_gr)
        if is_redirected_gr: fill_color = "FF0000" 
        should_be_in_pending_col = (is_pending and not is_pending_gr) or is_redirected_gr
        pending_row = {
            'Flow': processed['flow'], 'License Plat': processed['license_plat'], 'Driver': processed['driver_name'],
            'Faktur Batal/ Tolakan SO': processed['customer_name'] if is_batal else '',
            'Terkirim Sebagian': processed['customer_name'] if is_sebagian else '',
            'Pending': processed['customer_name'] if should_be_in_pending_col else '',
            'Reason': reason, 'Open Time': processed['open_time'], 'Close Time': processed['close_time'],
            'ETA': processed['eta'], 'ETD': processed['etd'], 'Actual Arrival': processed['actual_arrival'],
            'Actual Departure': processed['actual_departure'], 'Visit Time': processed['visit_time'],
            'Actual Visit Time': processed['actual_visit_time'], 'Customer ID': match.group(1) if match else 'N/A',
            'RO Sequence': ro_sequence, 'Real Sequence': real_sequence,
            'Temperature': ('DRY' if processed['driver_name'].startswith("'DRY'") else 'FRZ' if processed['driver_name'].startswith("'FRZ'") else 'N/A'),
            '_fill': fill_color 
        }
        if show_pending_gr:
            pending_row['Pending GR'] = processed['customer_name'] if is_pending_gr else ''
        pending_so_data.append(pending_row)
    if pending_so_data:
        df_pending = pd.DataFrame(pending_so_data)
    if not df_pending.empty:
        df_pending = df_pending.sort_values(by='Driver', ascending=True)
        if '_fill' in df_pending.columns:
            fill_values = df_pending['_fill']
        cols = list(df_pending.columns)
        if '_fill' in cols: cols.remove('_fill')
        if 'Pending GR' in cols and 'Pending' in cols:
            cols.insert(cols.index('Pending') + 1, cols.pop(cols.index('Pending GR')))
        if 'Reason' in cols:
            df_pending[' '] = '' 
            cols.insert(cols.index('Reason') + 1, ' ')
        df_pending = df_pending[cols] 
    
    # --- Logika 'Update Longlat' (Tidak diubah) ---
    update_longlat_data = []
    for task in tasks_data: 
        new_longlat = task.get('klikLokasiClient', '')
        old_longlat = task.get('longlat', '')
        if not new_longlat: continue
        beda_jarak = haversine_distance(old_longlat, new_longlat)
        title = task.get('title', '')
        match_id = re.search(r'C0\d{6,}', title)
        customer_id = match_id.group(0) if match_id else 'N/A'
        customer_name = title.split(' - ')[0].strip()
        parts = title.split(' - ')
        location_code_longlat = parts[-1].strip() if len(parts) > 2 else 'N/A'
        update_longlat_data.append({
            'Customer ID': customer_id, 'Customer Name': customer_name,
            'Location ID': location_code_longlat, 'New Longlat': new_longlat,
            'Beda Jarak (m)': beda_jarak
        })
    if update_longlat_data:
        df_longlat = pd.DataFrame(update_longlat_data)
        df_longlat = df_longlat.sort_values(by='Beda Jarak (m)', ascending=True)
    else:
        df_longlat = pd.DataFrame({"": ["Tidak Ada Update Longlat"]})

    # --- Simpan ke Excel ---
    lokasi_name = next((name for name, code in location_id.items() if code == LOKASI_FILTER), LOKASI_FILTER)
    selected_date_for_filename = dates["dmy"].replace("-", ".")
    base_name = f"Delivery Summary - {selected_date_for_filename} - {lokasi_name}"
    NAMA_FILE_OUTPUT = get_save_path(base_name)

    if not NAMA_FILE_OUTPUT: show_info_message("Dibatalkan", INFO_MESSAGES["CANCELED_BY_USER"]); return False

    try:
        simpan_excel(NAMA_FILE_OUTPUT, df_delivered, df_pending, df_ro_vs_real, df_longlat, fill_values, show_pending_gr)
        
        open_file_externally(NAMA_FILE_OUTPUT)
        return True
    except Exception as e: show_error_message("Gagal Menyimpan", ERROR_MESSAGES["UNKNOWN_ERROR"].format(error_detail=f"GAGAL MENYIMPAN FILE EXCEL: {e}\n\n{traceback.format_exc()}")); return False

# (Fungsi 'main' dan 'process_wrapper' tidak perlu diubah)

# (Fungsi panggil_api_dan_simpan sekarang memanggil ini)
def simpan_excel(NAMA_FILE_OUTPUT, df_delivered, df_pending, df_ro_vs_real, df_longlat, fill_values, show_pending_gr):
    """
    Menyimpan semua DataFrame ke dalam satu file Excel dengan beberapa sheet.
    (Versi Modifikasi: Mewarnai ETA/ETD jika Customer == 'HUB')
    """
    
    try:
        with pd.ExcelWriter(NAMA_FILE_OUTPUT, engine='openpyxl') as writer:
            
            # Sheet Total Delivered
            if not df_delivered.empty:
                 format_excel_sheet(writer, df_delivered, 'Total Delivered', centered_cols=['Total Outlet', 'Total Delivered'])
            else:
                 pd.DataFrame([{" ": "Tidak ada data kunjungan valid (filter Code 2)"}]) \
                   .to_excel(writer, sheet_name='Total Delivered', index=False)

            # Sheet Hasil Pending SO
            if not df_pending.empty:
                pending_centered_cols = ['Flow', 'Open Time', 'Close Time', 'ETA', 'ETD', 'Actual Arrival', 'Actual Departure', 'Visit Time', 'Actual Visit Time', 'Customer ID', 'RO Sequence', 'Real Sequence', 'Temperature']
                format_excel_sheet(writer, df_pending, 'Hasil Pending SO', centered_cols=pending_centered_cols, colored_cols={' ': "FFC0CB"})

                if fill_values is not None:
                    ws_pending = writer.sheets["Hasil Pending SO"]
                    bright_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    
                    pending_col_idx = None
                    for cell in ws_pending[1]: # Loop header (baris 1)
                        if cell.value == "Pending":
                            pending_col_idx = cell.column
                            break
                    
                    if pending_col_idx is not None:
                        for i, fill_val in enumerate(fill_values.values, start=2):
                            if pd.notna(fill_val):
                                ws_pending.cell(row=i, column=pending_col_idx).fill = bright_red_fill
            else:
                df_placeholder = pd.DataFrame({"Semua Pengiriman Sukses": []})
                df_placeholder.to_excel(writer, index=False, sheet_name='Hasil Pending SO')
                ws_pending = writer.sheets['Hasil Pending SO']
                ws_pending['A1'].font = Font(bold=True)
                ws_pending.column_dimensions['A'].width = len("Semua Pengiriman Sukses") + 5

            # Sheet Hasil RO vs Real
            ro_centered_cols = ['Flow', 'Status Delivery', 'Open Time', 'Close Time', 'ETA', 'Actual Arrival', 'ETD', 'Actual Departure', 'Visit Time', 'Actual Visit Time', 'RO Sequence', 'Real Sequence', 'Is Same Sequence']
            format_excel_sheet(writer, df_ro_vs_real, 'Hasil RO vs Real', centered_cols=ro_centered_cols)

            if "Hasil RO vs Real" in writer.sheets:
                ws_ro_vs_real = writer.sheets["Hasil RO vs Real"]
                RED_FONT = Font(color="FF0000")
                
                # Cari indeks kolom ETA, ETD, dan Customer (berbasis 1)
                col_idx = {}
                for cell in ws_ro_vs_real[1]: # Iterasi header
                    if cell.value in ["ETA", "ETD", "Customer"]:
                        col_idx[cell.value] = cell.column # Simpan nomor kolom
                
                # Pastikan semua kolom ditemukan
                if "Customer" in col_idx and ("ETA" in col_idx or "ETD" in col_idx):
                    
                    # Dapatkan nomor kolom
                    customer_col = col_idx["Customer"]
                    eta_col = col_idx.get("ETA") # Mungkin None jika kolom tdk ada
                    etd_col = col_idx.get("ETD") # Mungkin None jika kolom tdk ada
                
                    # Iterasi per baris, mulai dari baris 2
                    for row in ws_ro_vs_real.iter_rows(min_row=2):
                        
                        # Ambil sel Customer di baris ini
                        # (Nomor kolom berbasis 1, iter_rows berbasis 0, jadi kurangi 1)
                        customer_cell = row[customer_col - 1]
                        
                        # Cek apakah ini baris "HUB"
                        if customer_cell.value == "HUB":
                            
                            # Warnai sel Customer
                            customer_cell.font = RED_FONT
                            
                            # Warnai sel ETA di baris yang sama
                            if eta_col:
                                eta_cell = row[eta_col - 1]
                                eta_cell.font = RED_FONT
                            
                            # Warnai sel ETD di baris yang sama
                            if etd_col:
                                etd_cell = row[etd_col - 1]
                                etd_cell.font = RED_FONT

            # Sheet Update Longlat (Logika Code 1)
            if "Customer ID" in df_longlat.columns:
                longlat_centered_cols = ['Customer ID', 'Location ID', 'New Longlat', 'Beda Jarak (m)']
                format_excel_sheet(writer, df_longlat, 'Update Longlat', centered_cols=longlat_centered_cols)
            else:
                df_longlat.to_excel(writer, index=False, sheet_name='Update Longlat')

            # --- Penambahan Komentar (Tidak diubah) ---
            comment_author = "System" 
            if "Total Delivered" in writer.sheets:
                ws_ro = writer.sheets["Total Delivered"]
                comments_ro = { "Total Delivered": "Total Outlet - (Pending + Batal + Terima Sebagian)", }
                for cell in ws_ro[1]: 
                    if cell.value in comments_ro:
                        cell.comment = Comment(comments_ro[cell.value], comment_author)
            if "Hasil RO vs Real" in writer.sheets:
                ws_ro = writer.sheets["Hasil RO vs Real"]
                comments_ro = {
                    "RO Sequence": "Urutan berdasarkan hasil routing",
                    "Real Sequence": "Urutan kunjungan aktual di lapangan.",
                }
                for cell in ws_ro[1]: 
                    if cell.value in comments_ro:
                        cell.comment = Comment(comments_ro[cell.value], comment_author)
            if "Update Longlat" in writer.sheets:
                ws_longlat = writer.sheets["Update Longlat"]
                comments_longlat = { "Beda Jarak (m)": "Perhitungan jarak secara garis lurus antara koordinat lama dan baru." }
                for cell in ws_longlat[1]: 
                    if cell.value in comments_longlat:
                        cell.comment = Comment(comments_longlat[cell.value], comment_author)
            if fill_values is not None and fill_values.notna().any():
                if "Hasil Pending SO" in writer.sheets:
                    ws_pending_comment = writer.sheets["Hasil Pending SO"]
                    target_header = "Pending"
                    comment_text = 'Warna merah menandakan harusnya pilih "Pending" bukan "Pending GR"'
                    for cell in ws_pending_comment[1]: 
                        if cell.value == target_header:
                            cell.comment = Comment(comment_text, comment_author)
                            break 

        return True
        
    except Exception as e: 
        raise Exception(f"GAGAL MENYIMPAN FILE EXCEL: {e}\n\n{traceback.format_exc()}")

def main():
    def process_wrapper(dates, app_instance):
        def safe_close():
            if app_instance and app_instance.winfo_exists():
                app_instance.progress.stop()
                app_instance.destroy()
        try:
            panggil_api_dan_simpan(dates, app_instance)
        finally:
            if app_instance and app_instance.winfo_exists():
                app_instance.after(100, safe_close)

    create_date_picker_window("Delivery Summary", process_wrapper)

if __name__ == "__main__":
    main()