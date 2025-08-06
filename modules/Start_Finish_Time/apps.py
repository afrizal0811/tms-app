from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import openpyxl
import pandas as pd
import requests

# Impor fungsi bantuan dari shared_utils dan gui_utils
from utils.function import (
    get_save_path,
    load_config,
    load_constants,
    load_master_data,
    load_secret,
    open_file_externally,
    show_error_message,
    show_info_message
)
from utils.gui import create_date_picker_window
from utils.messages import ERROR_MESSAGES, INFO_MESSAGES

# =============================================================================
# BAGIAN 1: FUNGSI-FUNGSI BANTU (HELPER FUNCTIONS)
# =============================================================================

def extract_email_from_id(_id):
    parts = _id.split('_')
    return parts[1] if len(parts) > 1 else _id

def convert_to_jam(menit):
    jam = menit // 60
    sisa_menit = menit % 60
    return f"{jam}:{sisa_menit:02}"

def tambah_7_jam(waktu_str):
    waktu = datetime.strptime(waktu_str, "%Y-%m-%d %H:%M:%S")
    return (waktu + timedelta(hours=7)).strftime("%Y-%m-%d %H:%M:%S")

def simpan_file_excel(dataframe, lokasi_name, tanggal_str):
    kolom_buang = ['finish.lat', 'finish.lon', 'finish.notes']
    dataframe = dataframe.drop(columns=[col for col in kolom_buang if col in dataframe.columns])

    kolom_baru = {
        'Driver': 'Driver', 'startTime': 'Start Trip', 'finish.finishTime': 'Finish Trip',
        'trackedTime': 'Tracked Time', 'finish.totalDuration': 'Total Duration',
        'finish.totalDistance': 'Total Distance',
    }
    dataframe = dataframe.rename(columns=kolom_baru)

    for kolom in ['Start Trip', 'Finish Trip']:
        if kolom in dataframe.columns:
            dataframe[kolom] = pd.to_datetime(dataframe[kolom], errors='coerce')

    if 'Start Trip' in dataframe.columns:
        dataframe.insert(dataframe.columns.get_loc('Start Trip'), 'Start Date', dataframe['Start Trip'].dt.strftime('%d-%m-%Y'))
        dataframe['Start Trip'] = dataframe['Start Trip'].dt.strftime('%H:%M')

    if 'Finish Trip' in dataframe.columns:
        dataframe.insert(dataframe.columns.get_loc('Finish Trip'), 'Finish Date', dataframe['Finish Trip'].dt.strftime('%d-%m-%Y'))
        dataframe['Finish Trip'] = dataframe['Finish Trip'].dt.strftime('%H:%M')

    dataframe = dataframe.rename(columns={'Start Trip': 'Start Time', 'Finish Trip': 'Finish Time'})

    if 'Tracked Time' in dataframe.columns:
        dataframe['Tracked Time'] = dataframe['Tracked Time'].astype(str).apply(lambda x: f"'{x}" if pd.notna(x) and x != 'None' and x != '' else x)
    if 'Total Duration' in dataframe.columns:
        dataframe['Total Duration'] = dataframe['Total Duration'].astype(str).apply(lambda x: f"'{x}" if pd.notna(x) and x != 'None' and x != '' else x)

    if 'Total Distance' in dataframe.columns:
        dataframe['Total Distance'] = pd.to_numeric(dataframe['Total Distance'], errors='coerce')
        dataframe['Total Distance'] = dataframe['Total Distance'].round(2)
        dataframe['Total Distance'] = dataframe['Total Distance'].astype(object)
        dataframe.loc[dataframe['Total Distance'] == 0, 'Total Distance'] = ''

    urutan_kolom = [
        'Plat', 'Driver', 'Start Date', 'Start Time', 'Finish Date',
        'Finish Time', 'Tracked Time', 'Total Duration', 'Total Distance',
    ]
    urutan_kolom_final = [col for col in urutan_kolom if col in dataframe.columns]
    dataframe = dataframe[urutan_kolom_final]

    file_basename = f"Time Summary {lokasi_name} - {tanggal_str}"
    filename = get_save_path(file_basename)
    if not filename:
        show_info_message("Dibatalkan", INFO_MESSAGES["CANCELLED_BY_USER"])
        return

    dataframe.to_excel(filename, index=False)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    kolom_rata_tengah = ['Plat', 'Start Date', 'Start Time', 'Finish Date', 'Finish Time', 'Tracked Time', 'Total Duration', 'Total Distance']
    col_idx_map = {col_name: idx + 1 for idx, col_name in enumerate(dataframe.columns)}
    center_alignment = Alignment(horizontal='center', vertical='center')

    for col_name in kolom_rata_tengah:
        if col_name in col_idx_map:
            col_number = col_idx_map[col_name]
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_number).alignment = center_alignment

    for column_idx, column_name in enumerate(dataframe.columns):
        column_letter = get_column_letter(column_idx + 1)
        try:
            max_length = max(len(str(cell.value)) for cell in ws[column_letter] if cell.value is not None)
            adjusted_width = (max_length + 2)
            if adjusted_width > 0:
                ws.column_dimensions[column_letter].width = adjusted_width
        except ValueError:
            pass
            
    merah_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    start_date_col_idx = col_idx_map.get('Start Date')
    finish_date_col_idx = col_idx_map.get('Finish Date')
    if start_date_col_idx and finish_date_col_idx:
        for row in range(2, ws.max_row + 1):
            start_date_cell = ws.cell(row=row, column=start_date_col_idx)
            finish_date_cell = ws.cell(row=row, column=finish_date_col_idx)
            if start_date_cell.value and finish_date_cell.value and start_date_cell.value != finish_date_cell.value:
                start_date_cell.fill = merah_fill
                finish_date_cell.fill = merah_fill

    wb.save(filename)
    open_file_externally(filename)

# =============================================================================
# BAGIAN 2: FUNGSI PEMROSESAN UTAMA
# =============================================================================

def ambil_data(dates, app_instance=None):
    tanggal_str = dates["dmy"]

    if app_instance:
        app_instance.update_status("Mengambil data dari API...")

    config = load_config()
    constants = load_constants()
    secrets = load_secret()
    master_data = load_master_data()

    if not config or "lokasi" not in config:
        show_error_message("Error Konfigurasi", ERROR_MESSAGES["CONFIG_FILE_ERROR"])
        return False
    if not constants:
        show_error_message("Error Konstanta", ERROR_MESSAGES["CONSTANT_FILE_ERROR"])
        return False
    if not secrets:
        show_error_message("Error Rahasia", ERROR_MESSAGES["SECRET_FILE_ERROR"])
        return False
    if not master_data:
        show_error_message("Error Master", ERROR_MESSAGES["MASTER_DATA_MISSING"])
        return False

    api_token = secrets.get('token')
    lokasi_code = config.get('lokasi')
    hub_ids = master_data.get('hub_ids', {})
    lokasi_mapping = constants.get('lokasi_mapping', {})

    if not api_token:
        show_error_message("Error Token API", ERROR_MESSAGES["API_TOKEN_MISSING"])
        return False
    if not lokasi_code:
        show_error_message("Error Konfigurasi", ERROR_MESSAGES["LOCATION_CODE_MISSING"])
        return False
    if lokasi_code not in hub_ids:
        show_error_message("Error Hub ID", ERROR_MESSAGES["HUB_ID_MISSING"])
        return False

    lokasi_name = next((name for name, code in lokasi_mapping.items() if code == lokasi_code), lokasi_code)

    tanggal_obj = datetime.strptime(tanggal_str, "%d-%m-%Y")
    tanggal_input = tanggal_obj.strftime("%Y-%m-%d")
    tanggal_from = (tanggal_obj - timedelta(days=1)).strftime("%Y-%m-%d")

    base_url = constants.get('base_url')
    url = f"{base_url}/location-histories"
    params = {
        "limit": 150, "startFinish": "true", "fields": "finish,startTime",
        "timeTo": f"{tanggal_input} 23:59:59", "timeFrom": f"{tanggal_from} 00:00:00",
        "timeBy": "createdTime"
    }
    headers = {"Authorization": f"Bearer {api_token}"}

    try:
        response = requests.get(url, params=params, headers=headers, timeout=30)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        show_error_message("API Error", ERROR_MESSAGES["API_REQUEST_FAILED"].format(error_detail=e))
        return

    items = response.json().get("tasks", {}).get("data", [])
    if not items:
        show_error_message("Error", ERROR_MESSAGES["DATA_NOT_FOUND"])
        return

    filtered_items = []
    for item in items:
        if item.get("trackedTime") not in [None, 0] and item.get("trackedTime", 0) >= 10 and item.get("finish", {}).get("totalDistance", float('inf')) > 5:
            item_copy = item.copy()
            item_copy["_id"] = extract_email_from_id(item_copy["_id"])
            item_copy["startTime"] = tambah_7_jam(item_copy["startTime"])
            if item_copy.get("finish"):
                item_copy["finish"]["finishTime"] = tambah_7_jam(item_copy["finish"]["finishTime"])
                if "totalDuration" in item_copy["finish"]:
                    item_copy["finish"]["totalDuration"] = convert_to_jam(item_copy["finish"]["totalDuration"])
            if "trackedTime" in item_copy:
                item_copy["trackedTime"] = convert_to_jam(item_copy["trackedTime"])
            filtered_items.append(item_copy)

    df_api_data = pd.json_normalize(filtered_items)
    if df_api_data.empty:
        show_error_message("Gagal", ERROR_MESSAGES["DATA_NOT_FOUND"])
        return

    df_api_data.rename(columns={"_id": "Email"}, inplace=True)
    if 'startTime' in df_api_data.columns:
        df_api_data['startTime'] = pd.to_datetime(df_api_data['startTime'])
        df_api_data = df_api_data[df_api_data['startTime'].dt.strftime("%Y-%m-%d") == tanggal_input]

    try:
        df_api_data = df_api_data[df_api_data['Email'].str.contains(lokasi_code, na=False, case=False)]
        if df_api_data.empty:
            show_error_message("Error", ERROR_MESSAGES["DATA_NOT_FOUND"])
            return

        mapping_df = master_data["df"]
        mapping_df_filtered = mapping_df[mapping_df['Email'].str.contains(lokasi_code, na=False, case=False)]
        df_merged = df_api_data.merge(mapping_df_filtered[['Email', 'Driver', 'Plat']], on='Email', how='left')

        df_merged['Driver'] = df_merged['Driver'].fillna(df_merged['Email'])
        df_merged['Plat'] = df_merged['Plat'].fillna('')
        df_merged.drop(columns=['Email'], inplace=True)
        df_merged = df_merged.sort_values(by='Driver', ascending=True)

    except Exception as e:
        show_error_message("Error", ERROR_MESSAGES["UNKNOWN_ERROR"].format(error_detail=e))
        return

    existing_drivers = df_merged['Driver'].unique().tolist()
    missing_rows_df = mapping_df_filtered[~mapping_df_filtered['Driver'].isin(existing_drivers)].copy()
    if not missing_rows_df.empty:
        missing_drivers_df_processed = missing_rows_df[['Driver', 'Plat']].copy()
        for col in df_merged.columns:
            if col not in missing_drivers_df_processed.columns:
                missing_drivers_df_processed[col] = ""
        missing_drivers_df_processed = missing_drivers_df_processed[df_merged.columns]
        final_df = pd.concat([df_merged, missing_drivers_df_processed], ignore_index=True)
    else:
        final_df = df_merged.copy()

    def durasi_ke_menit(durasi_str):
        try:
            if isinstance(durasi_str, str) and ':' in durasi_str:
                jam, menit = map(int, durasi_str.split(':'))
                return jam * 60 + menit
        except (ValueError, AttributeError):
            pass
        return 0

    final_df['finish.totalDuration_menit'] = final_df['finish.totalDuration'].apply(durasi_ke_menit)
    final_df['finish.totalDistance'] = pd.to_numeric(final_df['finish.totalDistance'], errors='coerce').fillna(0)
    final_df = final_df.sort_values(by='Driver')

    def filter_duplikat(grup):
        idx_to_drop = set()
        for i in range(len(grup)):
            for j in range(len(grup)):
                if i != j:
                    row_i = grup.iloc[i]
                    row_j = grup.iloc[j]
                    if (row_i['finish.totalDuration_menit'] < row_j['finish.totalDuration_menit'] and
                        row_i['finish.totalDistance'] < row_j['finish.totalDistance']):
                        idx_to_drop.add(grup.index[i])
        return grup.drop(list(idx_to_drop))

    non_empty_df = final_df[final_df['finish.totalDuration_menit'] > 0]
    groups = [filter_duplikat(group) for _, group in non_empty_df.groupby('Driver')]
    filtered_df_final = pd.concat(groups, ignore_index=True) if groups else pd.DataFrame()

    kosong_df = final_df[final_df['finish.totalDuration_menit'] == 0]
    final_df = pd.concat([filtered_df_final, kosong_df], ignore_index=True)
    final_df.drop(columns=['finish.totalDuration_menit'], inplace=True)
    final_df = final_df.sort_values(by='Driver', ascending=True)

    tanggal_format_titik = tanggal_str.replace('-', '.')
    simpan_file_excel(final_df, lokasi_name, tanggal_format_titik)

    if app_instance:
        app_instance.update_status("Proses selesai.")

# =============================================================================
# BAGIAN 3: FUNGSI GUI DAN EKSEKUSI
# =============================================================================

def main():
    config = load_config()
    if not config or not config.get("lokasi"):
        show_error_message("Setup Awal", ERROR_MESSAGES["LOCATION_CODE_MISSING"])
        return

    def process_wrapper(dates, app_instance):
        ambil_data(dates, app_instance)

    create_date_picker_window("Start-Finish Time", process_wrapper)

if __name__ == "__main__":
    main()
