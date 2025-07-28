# modules/Delivery_Summary/apps.py (KODE BARU)

# -*- coding: utf-8 -*-
"""
Skrip Gabungan untuk memproses laporan "RO vs Real" dan "Pending SO".
"""

# 1. Impor yang dibutuhkan sudah dirapikan
import pandas as pd
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

# 2. Impor semua fungsi bantuan dari shared_utils
from ..shared_utils import load_config, load_master_data, get_save_path, open_file_externally


# =============================================================================
# BAGIAN 1: FUNGSI-FUNGSI BANTU (HELPER FUNCTIONS)
# =============================================================================

# SEMUA FUNGSI LAMA (load_config, baca_master_driver, get_save_path, open_file)
# SUDAH DIHAPUS DARI SINI KARENA KITA MENGGUNAKAN VERSI DARI shared_utils.py


# Fungsi-fungsi yang spesifik untuk modul ini tetap ada di sini
def apply_styles_and_formatting(writer):
    """
    Menerapkan semua styling (alignment, warna, auto-size) ke semua sheet.
    """
    workbook = writer.book
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    cols_to_center = [
        'Open Time', 'Close Time', 'ETA', 'ETD', 'Actual Arrival', 
        'Actual Departure', 'Visit Time', 'Actual Visit Time', 
        'Customer ID', 'ET Sequence', 'Real Sequence', 'Temperature',
        'Total Visit', 'Total Delivered', 'Status Delivery', 'Is Same Sequence'
    ]

    for sheet_name in workbook.sheetnames:
        worksheet = writer.sheets[sheet_name]
        header_map = {cell.value: cell.column for cell in worksheet[1]}
        for col_name, col_idx in header_map.items():
            col_letter = get_column_letter(col_idx)
            align = center_align if col_name in cols_to_center else left_align
            for cell in worksheet[col_letter]:
                cell.alignment = align
                if col_name == ' ':
                    cell.fill = red_fill
        if ' ' in header_map:
            sep_col_idx = header_map[' ']
            worksheet.cell(row=1, column=sep_col_idx).value = ""
        for column_cells in worksheet.columns:
            # Menghindari error jika kolom kosong
            try:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                adjusted_width = max_length + 2
                if adjusted_width > 50: adjusted_width = 50
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
            except ValueError:
                # Lewati jika kolom benar-benar kosong
                pass

def convert_datetime_column(df, column_name, target_format='%H:%M'):
    """Mengonversi kolom datetime ke format string waktu."""
    def convert(val):
        if pd.isna(val) or val == '': return ''
        try:
            if isinstance(val, datetime): dt = val
            elif 'T' in str(val): dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
            else: dt = pd.to_datetime(val)
            return dt.strftime(target_format)
        except Exception: return val
    df[column_name] = df[column_name].apply(convert)
    return df

def insert_blank_rows(df, column):
    """Menyisipkan baris kosong ketika nilai di kolom tertentu berubah."""
    if df.empty: return df
    df.loc[:, column] = df[column].fillna('')
    df = df.sort_values(by=column, ascending=True)
    new_rows = []
    prev_value = None
    for _, row in df.iterrows():
        current_value = row[column]
        if prev_value is not None and current_value != prev_value:
            new_rows.append(pd.Series([None]*len(df.columns), index=df.columns))
        new_rows.append(row)
        prev_value = current_value
    return pd.DataFrame(new_rows).reset_index(drop=True)

def calculate_actual_visit(start, end):
    """Menghitung durasi kunjungan dalam menit."""
    if start == '' or end == '' or pd.isna(start) or pd.isna(end): return ""
    try:
        t1 = datetime.strptime(str(start), "%H:%M")
        t2 = datetime.strptime(str(end), "%H:%M")
        delta = (t2 - t1).total_seconds()
        if delta < 0: delta += 86400
        return int(delta // 60)
    except (ValueError, TypeError): return ""


# =============================================================================
# BAGIAN 2: FUNGSI-FUNGSI PEMROSESAN UTAMA (TIDAK ADA PERUBAHAN DI SINI)
# =============================================================================
def process_total_delivered(df, master_driver_df):
    """Membuat ringkasan jumlah visit dan delivery, membiarkan total kosong untuk driver tanpa data."""
    master_summary = master_driver_df[['Driver', 'Plat']].copy()
    master_summary.rename(columns={'Plat': 'License Plat'}, inplace=True)
    master_summary.drop_duplicates(subset=['Driver'], inplace=True)

    if 'assignee' in df.columns and 'label' in df.columns:
        df_proc = df.copy()
        email_to_name = dict(zip(master_driver_df['Email'], master_driver_df['Driver']))
        df_proc['Driver'] = df_proc['assignee'].str.lower().map(email_to_name)
        df_proc.dropna(subset=['Driver'], inplace=True)

        visit_counts = df_proc['Driver'].value_counts().reset_index()
        visit_counts.columns = ['Driver', 'Total Visit']

        delivered_df = df_proc[df_proc['label'].str.upper() == 'SUKSES'].copy()
        delivered_counts = delivered_df['Driver'].value_counts().reset_index()
        delivered_counts.columns = ['Driver', 'Total Delivered']

        final_df = pd.merge(master_summary, visit_counts, on='Driver', how='left')
        final_df = pd.merge(final_df, delivered_counts, on='Driver', how='left')
    else:
        final_df = master_summary
        final_df['Total Visit'] = pd.NA
        final_df['Total Delivered'] = pd.NA

    final_df['License Plat'] = final_df['License Plat'].fillna('-')
    final_df = final_df.sort_values(by='Driver', ascending=True).reset_index(drop=True)
    
    df_final = final_df[['License Plat', 'Driver', 'Total Visit', 'Total Delivered']]
    
    if 'Total Visit' in df_final.columns:
        df_final['Total Visit'] = df_final['Total Visit'].astype('Int64')
    if 'Total Delivered' in df_final.columns:
        df_final['Total Delivered'] = df_final['Total Delivered'].astype('Int64')
        
    return df_final

def process_ro_vs_real(df, master_driver_df):
    df_proc = df.copy()
    email_to_name = dict(zip(master_driver_df['Email'], master_driver_df['Driver']))
    email_to_plat = dict(zip(master_driver_df['Email'], master_driver_df['Plat']))
    if 'assignee' in df_proc.columns:
        df_proc['assignee_email'] = df_proc['assignee'].str.lower()
        df_proc['assignee'] = df_proc['assignee_email'].map(email_to_name).fillna(df_proc['assignee'])
    if 'assignedVehicle' in df_proc.columns and 'assignee_email' in df_proc.columns:
        df_proc['assignedVehicle'] = df_proc.apply(lambda row: email_to_plat.get(row['assignee_email'], row['assignedVehicle']) if pd.isna(row['assignedVehicle']) or str(row['assignedVehicle']).strip() == '-' else row['assignedVehicle'], axis=1)
    time_columns = ['Klik Jika Anda Sudah Sampai', 'doneTime']
    for col in time_columns:
        if col in df_proc.columns: df_proc = convert_datetime_column(df_proc, col)
    if 'Klik Jika Anda Sudah Sampai' in df_proc.columns and 'doneTime' in df_proc.columns:
        df_proc['Actual Visit Time'] = df_proc.apply(lambda row: calculate_actual_visit(row['Klik Jika Anda Sudah Sampai'], row['doneTime']), axis=1)
    if 'doneTime' in df_proc.columns and 'assignee' in df_proc.columns:
        df_proc['doneTime_parsed'] = pd.to_datetime(df_proc['doneTime'], format='%H:%M', errors='coerce')
        df_proc['Real Seq'] = df_proc.groupby('assignee')['doneTime_parsed'].rank(method='dense').astype('Int64')
        df_proc.drop(columns=['doneTime_parsed'], inplace=True)

    rename_dict = {
        'assignedVehicle': 'License Plat', 'assignee': 'Driver', 'title': 'Customer',
        'label': 'Status Delivery', 'Klik Jika Anda Sudah Sampai': 'Actual Arrival',
        'doneTime': 'Actual Departure', 'routePlannedOrder': 'ET Sequence',
        'Real Seq': 'Real Sequence'
    }
    df_proc.rename(columns=rename_dict, inplace=True)
    df_proc = df_proc.loc[:,~df_proc.columns.duplicated()]

    if 'ET Sequence' in df_proc.columns and 'Real Sequence' in df_proc.columns:
        df_proc['Is Same Sequence'] = (pd.to_numeric(df_proc['ET Sequence'], errors='coerce') == pd.to_numeric(df_proc['Real Sequence'], errors='coerce'))
        df_proc['Is Same Sequence'] = df_proc['Is Same Sequence'].map({True: 'SAMA', False: 'TIDAK SAMA', pd.NA: ''})

    desired_columns = [
        'License Plat', 'Driver', 'Customer', 'Status Delivery', 'Open Time',
        'Close Time', 'Actual Arrival', 'Actual Departure', 'Visit Time',
        'Actual Visit Time', 'ET Sequence', 'Real Sequence', 'Is Same Sequence'
    ]
    final_cols = [col for col in desired_columns if col in df_proc.columns]
    df_final = df_proc[final_cols].copy()

    if 'Driver' in df_final.columns and 'Real Sequence' in df_final.columns and not df_final.empty:
        df_final['Real Sequence'] = pd.to_numeric(df_final['Real Sequence'], errors='coerce')
        df_final.sort_values(by=['Driver', 'Real Sequence'], ascending=True, inplace=True)
        all_drivers_data = []
        for _, group in df_final.groupby('Driver', sort=False):
            all_drivers_data.append(group)
            blank_row = pd.DataFrame([[None] * len(df_final.columns)], columns=df_final.columns)
            all_drivers_data.append(blank_row)
        if all_drivers_data:
            df_final = pd.concat(all_drivers_data[:-1], ignore_index=True)

    return df_final

def process_pending_so(df, master_driver_df):
    df_proc = df.copy()
    email_to_name = dict(zip(master_driver_df['Email'], master_driver_df['Driver']))
    if 'assignee' in df_proc.columns:
        df_proc['Driver'] = df_proc['assignee'].str.lower().map(email_to_name).fillna(df_proc['assignee'])
    if 'Klik Jika Anda Sudah Sampai' in df_proc.columns and 'assignee' in df_proc.columns:
        df_proc['arrival_datetime_temp'] = pd.to_datetime(df_proc['Klik Jika Anda Sudah Sampai'], errors='coerce')
        df_proc['Real Seq'] = df_proc.groupby('assignee')['arrival_datetime_temp'].rank(method='dense').astype('Int64')
        df_proc.drop(columns=['arrival_datetime_temp'], inplace=True)
    if 'label' not in df_proc.columns: return None
    status_to_filter = ['BATAL', 'PENDING', 'TERIMA SEBAGIAN']
    df_filtered = df_proc[df_proc['label'].isin(status_to_filter)].copy()
    if df_filtered.empty: return None 
    time_cols_to_convert = ['Klik Jika Anda Sudah Sampai', 'doneTime', 'eta', 'etd']
    for col in time_cols_to_convert:
        if col in df_filtered.columns: df_filtered = convert_datetime_column(df_filtered, col)
    if 'Klik Jika Anda Sudah Sampai' in df_filtered.columns and 'doneTime' in df_filtered.columns:
         df_filtered['Actual Visit Time'] = df_filtered.apply(lambda row: calculate_actual_visit(row['Klik Jika Anda Sudah Sampai'], row['doneTime']), axis=1)
    def extract_customer_id(title):
        if not isinstance(title, str): return ''
        try:
            parts = title.split('-'); return parts[1].strip() if len(parts) > 1 else ''
        except IndexError: return ''
    if 'title' in df_filtered.columns: df_filtered['Customer ID'] = df_filtered['title'].apply(extract_customer_id)
    if 'Driver' in df_filtered.columns: df_filtered['Temperature'] = df_filtered['Driver'].str.split(' ').str[0].str.replace("'", "")
    def get_reason(row):
        status = row.get('label', ''); return row.get('Alasan Batal', '') if status in ['PENDING', 'BATAL'] else (row.get('Alasan Tolakan', '') if status == 'TERIMA SEBAGIAN' else '')
    df_filtered['Reason'] = df_filtered.apply(get_reason, axis=1)
    def assign_faktur_by_title(row):
        status, title_val = row.get('label', ''), row.get('title', ''); return (title_val, '', '') if status == 'BATAL' else (('', title_val, '') if status == 'TERIMA SEBAGIAN' else (('', '', title_val) if status == 'PENDING' else ('', '', '')))
    (df_filtered['Faktur Batal/ Tolakan SO'], df_filtered['Terkirim Sebagian'], df_filtered['Pending']) = zip(*df_filtered.apply(assign_faktur_by_title, axis=1))
    
    kolom_final = ['assignedVehicle', 'Driver', 'Faktur Batal/ Tolakan SO', 'Terkirim Sebagian', 'Pending', 'Reason', 'Open Time', 'Close Time', 'eta', 'etd', 'Klik Jika Anda Sudah Sampai', 'doneTime', 'Visit Time', 'Actual Visit Time', 'Customer ID', 'routePlannedOrder', 'Real Seq', 'Temperature']
    rename_kolom = {'assignedVehicle': 'License Plat', 'eta': 'ETA', 'etd': 'ETD', 'Klik Jika Anda Sudah Sampai': 'Actual Arrival', 'doneTime': 'Actual Departure', 'routePlannedOrder': 'ET Sequence', 'Real Seq': 'Real Sequence'}
    cols_to_select = [col for col in kolom_final if col in df_filtered.columns]
    df_final = df_filtered[cols_to_select].copy()
    
    if 'Reason' in df_final.columns:
        reason_loc = df_final.columns.get_loc('Reason')
        df_final.insert(loc=reason_loc + 1, column=' ', value='')
    
    df_final.rename(columns=rename_kolom, inplace=True)
    if 'Driver' in df_final.columns: df_final = df_final.sort_values(by='Driver', ascending=True).reset_index(drop=True)
    return df_final


# =============================================================================
# BAGIAN 3: FUNGSI EKSEKUSI UTAMA
# =============================================================================

def main():
    # 3. Ganti pemanggilan fungsi dengan versi dari shared_utils
    config = load_config()
        
    if config and "lokasi" in config:
        lokasi = config["lokasi"]
    else:
        messagebox.showwarning("Dibatalkan", "Pilih lokasi cabang!")
        return

    messagebox.showinfo("Informasi", "Pilih Export Task")
    
    input_file = filedialog.askopenfilename(title="Pilih File Excel yang Akan Diproses", filetypes=[("Excel Files", "*.xlsx *.xls")])
    
    if not input_file:
        messagebox.showwarning("Proses Gagal", "Proses Dibatalkan")
        return

    try:
        df_original = pd.read_excel(input_file)
        
        # Blok Validasi File (TIDAK BERUBAH)
        required_columns = ['assignedVehicle', 'assignee', 'Alasan Tidak Bisa Dikunjungi', 'Alasan Batal','Open Time', 'Close Time', 'eta', 'etd', 'Klik Jika Anda Sudah Sampai', 'doneTime', 'Visit Time', 'routePlannedOrder']
        missing_columns = [col for col in required_columns if col not in df_original.columns]
        if missing_columns:
            messagebox.showerror("Proses Gagal", f"File tidak valid!\n\nUpload file Export Task dengan benar!")
            return
            
        email_prefixes = df_original["assignee"].dropna().astype(str).str.extract(r'kendaraan\.([^.@]+)', expand=False)
        email_prefixes = email_prefixes.dropna().str.lower().unique()
        if not any(lokasi.lower() in prefix for prefix in email_prefixes):
            messagebox.showerror("Proses Gagal", f"Lokasi cabang tidak valid!\n\nLokasi cabang tidak sesuai dengan file Export Task!")
            return

        results_to_save = {}
        
        # Panggil fungsi terpusat untuk memuat data master
        master_df = load_master_data(lokasi)
        if master_df is None:
            return

        result_total = process_total_delivered(df_original, master_df)
        if result_total is not None and not result_total.empty:
            results_to_save['Total Delivered'] = result_total

        result_pending = process_pending_so(df_original, master_df)
        if result_pending is not None and not result_pending.empty:
            results_to_save['Hasil Pending SO'] = result_pending
        
        result_ro = process_ro_vs_real(df_original, master_df)
        if result_ro is not None and not result_ro.empty:
            results_to_save['Hasil RO vs Real'] = result_ro
                
        if not results_to_save:
            messagebox.showerror("Proses Gagal", "File tidak valid atau tidak ada data yang relevan untuk diproses.")
            return
        
        # Panggil fungsi terpusat untuk mendapatkan path penyimpanan
        save_file_path = get_save_path("Delivery Summary")
        
        if not save_file_path:
            messagebox.showwarning("Proses Gagal", "Proses Dibatalkan")
            return
            
        with pd.ExcelWriter(save_file_path, engine='openpyxl') as writer:
            sheet_order = ['Total Delivered', 'Hasil Pending SO', 'Hasil RO vs Real']
            for sheet_name in sheet_order:
                if sheet_name in results_to_save:
                    results_to_save[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            
            apply_styles_and_formatting(writer)
        
        # Panggil fungsi terpusat untuk membuka file
        open_file_externally(save_file_path)
        
    except Exception as e:
        messagebox.showerror("Terjadi Error", f"Sebuah kesalahan tak terduga terjadi:\n{e}")

if __name__ == "__main__":
    main()