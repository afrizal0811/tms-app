import requests
import re 
import tkinter as tk
from tkinter import ttk 
from datetime import datetime, timedelta
import os
import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, PatternFill
from utils.function import load_config, load_secret, load_constants, show_error_message, load_master_data
from utils.messages import ERROR_MESSAGES
from utils.gui import create_date_picker_window
from utils.api_handler import handle_requests_error

# Batasan untuk tab/tombol kendaraan per halaman
VEHICLES_PER_PAGE = 10 

# =============================================================================
# FUNGSI EXPORT KE EXCEL 
# =============================================================================
def export_to_excel(all_vehicle_data, date_str, lokasi_cabang):
    if not all_vehicle_data:
        messagebox.showwarning("Tidak Ada Data", "Tidak ada data untuk diekspor.")
        return

    folder_path = filedialog.askdirectory(title="Pilih Folder untuk Menyimpan Laporan")
    
    if not folder_path:
        return

    try:
        base_filename = f"Estimasi Delivery - {date_str} - {lokasi_cabang}.xlsx"
        file_path = os.path.join(folder_path, base_filename)

        counter = 1
        while os.path.exists(file_path):
            name, ext = os.path.splitext(base_filename)
            file_path = os.path.join(folder_path, f"{name} ({counter}){ext}")
            counter += 1

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            red_font = Font(color="FF0000")
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            
            fill_target_columns = ['SO', 'Jam Buka', 'Jam Tutup', 'Estimasi Sampai', 'Estimasi Berangkat']

            for vehicle_data in all_vehicle_data:
                stop_details = vehicle_data.get('stopDetails', [])
                vehicle_name = vehicle_data.get('vehicleName', 'Sheet')
                
                export_rows = []
                min_order, max_order = (None, None)
                if stop_details:
                    orders = [s.get('order', 999) for s in stop_details]
                    min_order, max_order = (min(orders), max(orders))

                for stop in sorted(stop_details, key=lambda x: x.get('order', 999)):
                    visit_name = stop.get('visitName', 'N/A')
                    jam_buka = stop.get('startTime', 'N/A')
                    jam_tutup = stop.get('endTime', 'N/A')
                    eta_val = stop.get('eta', 'N/A')
                    etd_val = stop.get('etd', 'N/A')
                    so_numbers = stop.get('soNumbers', '')

                    if visit_name == 'HUB':
                        jam_buka = ""
                        jam_tutup = ""
                        current_order = stop.get('order')
                        if current_order == min_order: eta_val = ""
                        elif current_order == max_order: etd_val = ""
                    
                    export_rows.append({
                        'No.': stop.get('order'),
                        'Outlet': visit_name,
                        'SO': so_numbers,
                        'Jam Buka': jam_buka,
                        'Jam Tutup': jam_tutup,
                        'Estimasi Sampai': f"{eta_val.split(':')[0]}:{eta_val.split(':')[1]}" if ':' in str(eta_val) else eta_val,
                        'Estimasi Berangkat': f"{etd_val.split(':')[0]}:{etd_val.split(':')[1]}" if ':' in str(etd_val) else etd_val
                    })

                if export_rows:
                    df = pd.DataFrame(export_rows)
                    safe_sheet_name = re.sub(r'[\\/*?:"<>|]', "", vehicle_name)[:31]
                    df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                    worksheet = writer.sheets[safe_sheet_name]
                    for col in worksheet.columns:
                        max_length = 0
                        column_letter = col[0].column_letter
                        header_length = len(str(worksheet[f"{column_letter}1"].value))
                        max_length = header_length
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except: pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
                        if row_data.Outlet == 'HUB':
                            for col_idx in range(1, df.shape[1] + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.font = red_font
                                header_name = df.columns[col_idx-1]
                                if header_name in fill_target_columns and not cell.value:
                                    cell.fill = black_fill
        
        os.startfile(file_path)

    except Exception as e:
        messagebox.showerror("Error Ekspor", f"Terjadi kesalahan saat mengekspor data:\n{e}")

# =============================================================================
# FUNGSI UTILITY DATA & REGEX
# =============================================================================
def extract_outlet_name(visit_name):
    if not visit_name: return "N/A"
    separator = ' - '
    index = visit_name.find(separator)
    if index != -1: return visit_name[:index].strip()
    return visit_name.strip()

def get_hub_id():
    config = load_config()
    lokasi_code = config.get('lokasi') if config else None
    if not config or not lokasi_code:
        show_error_message("Error Konfigurasi", ERROR_MESSAGES.get("LOCATION_CODE_MISSING", "Kode lokasi hilang."))
        return None
    master_data = load_master_data() 
    hub_ids_map = master_data.get('hub_ids') if master_data else None
    hub_id = hub_ids_map.get(lokasi_code) if hub_ids_map else None
    if not hub_id:
        show_error_message("Error Hub ID", f"Hub ID untuk lokasi '{lokasi_code}' tidak ditemukan di master data.")
        return None
    return hub_id

def extract_customer_and_location(visit_name):
    if not visit_name: return "", ""
    cust_code = re.search(r'(C0\d+)', visit_name)
    loc_code = re.search(r'(MAIN|SHIPTO|LOC\d+)', visit_name)
    return cust_code.group(1) if cust_code else "", loc_code.group(1) if loc_code else ""

# =============================================================================
# FUNGSI UTILITY GUI
# =============================================================================
def create_vehicle_tab(notebook, vehicle_data):
    vehicle_name = vehicle_data['vehicleName']
    stop_details = vehicle_data['stopDetails'] 
    num_trips = vehicle_data['numTrips'] 
    
    tab_frame_container = ttk.Frame(notebook, padding="10")
    notebook.add(tab_frame_container, text=vehicle_name)
    
    summary_frame = tk.Frame(tab_frame_container, padx=10, pady=5, relief="raised", bd=1)
    summary_frame.pack(fill='x', pady=(0, 10))
    ttk.Label(summary_frame, text=f"Total Stop Pelanggan (Asli):", font=("Arial", 10)).pack(side=tk.LEFT, padx=10)
    ttk.Label(summary_frame, text=f"{num_trips}", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
    
    columns = ("No.", "Outlet", "SO", "Jam Buka", "Jam Tutup", "Estimasi Sampai", "Estimasi Berangkat")
    tree = ttk.Treeview(tab_frame_container, columns=columns, show='headings')
    
    vsb = ttk.Scrollbar(tab_frame_container, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(tab_frame_container, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    
    vsb.pack(side='right', fill='y')
    hsb.pack(side='bottom', fill='x')
    tree.pack(expand=True, fill='both')
    
    tree.tag_configure('hub_style', foreground='red')
    
    tree.column("No.", width=40, anchor='center', stretch=False)
    tree.column("Outlet", width=250)
    tree.column("SO", width=200)
    tree.column("Jam Buka", width=100, anchor='center', stretch=False)
    tree.column("Jam Tutup", width=100, anchor='center', stretch=False)
    tree.column("Estimasi Sampai", width=120, anchor='center', stretch=False)
    tree.column("Estimasi Berangkat", width=120, anchor='center', stretch=False)
    
    tree.heading("No.", text="No.")
    tree.heading("Outlet", text="Outlet")
    tree.heading("SO", text="SO")
    tree.heading("Jam Buka", text="Jam Buka")
    tree.heading("Jam Tutup", text="Jam Tutup")
    tree.heading("Estimasi Sampai", text="Estimasi Sampai")
    tree.heading("Estimasi Berangkat", text="Estimasi Berangkat")

    min_order, max_order = None, None
    if stop_details:
        orders = [s['order'] for s in stop_details]
        min_order = min(orders)
        max_order = max(orders)

    for stop in sorted(stop_details, key=lambda x: x['order']):
        visit_name = stop.get('visitName', 'N/A')
        jam_buka = stop.get('startTime', 'N/A')
        jam_tutup = stop.get('endTime', 'N/A')
        eta_val = stop.get('eta', 'N/A')
        etd_val = stop.get('etd', 'N/A')
        so_numbers = stop.get('soNumbers', '')
        tags_to_apply = ()
        is_hub = (visit_name == 'HUB')
        if is_hub:
            tags_to_apply = ('hub_style',)
            jam_buka = ""
            jam_tutup = ""
            current_order = stop.get('order')
            if current_order == min_order: eta_val = ""
            elif current_order == max_order: etd_val = ""
            
        eta_short = f"{eta_val.split(':')[0]}:{eta_val.split(':')[1]}" if ':' in str(eta_val) else eta_val
        etd_short = f"{etd_val.split(':')[0]}:{etd_val.split(':')[1]}" if ':' in str(etd_val) else etd_val
        
        tree.insert('', 'end', values=(
            stop.get('order', 999), 
            visit_name, 
            so_numbers,
            jam_buka, 
            jam_tutup, 
            eta_short,
            etd_short
        ), tags=tags_to_apply)

def create_summary_tab(notebook, vehicle_data):
    num_vehicles = len(vehicle_data)
    summary_frame = ttk.Frame(notebook, padding="20")
    notebook.add(summary_frame, text="Rangkuman") 
    center_frame = tk.Frame(summary_frame)
    center_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
    tk.Label(center_frame, text="REKAPITULASI ESTIMASI DELIVERY", font=("Arial", 16, "bold"), pady=15).pack()
    tk.Label(center_frame, text=f"Total {num_vehicles} Kendaraan Ditemukan", font=("Arial", 14, "bold"), fg="blue").pack(pady=5)
    tk.Label(center_frame, text="Gunakan pencarian global di atas untuk memfilter kendaraan.", font=("Arial", 10), fg="gray").pack(pady=5)

def display_result_gui(parent_instance, parsed_data, date_str, lokasi_cabang):
    if not parent_instance.winfo_exists(): return
    result_window = tk.Toplevel(parent_instance) 
    parent_instance.withdraw() 
    result_window.title(f"Hasil Estimasi Delivery - {date_str}")
    window_width, window_height = 1000, 700
    screen_width, screen_height = result_window.winfo_screenwidth(), result_window.winfo_screenheight()
    center_x = int(screen_width/2 - window_width / 2)
    center_y = int(screen_height/2 - window_height / 2)
    result_window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    main_container = tk.Frame(result_window); main_container.pack(expand=True, fill='both', padx=10, pady=10)
    
    top_control_frame = tk.Frame(main_container, pady=5); top_control_frame.pack(fill='x', anchor='n') 
    search_frame = ttk.Frame(top_control_frame)
    search_frame.pack(pady=(0, 5))
    ttk.Label(search_frame, text="Cari (Kendaraan / SO / Outlet):", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
    search_entry = ttk.Entry(search_frame, width=40)
    search_entry.pack(side=tk.LEFT, fill='x', expand=True)
    
    pagination_control_frame = tk.Frame(top_control_frame); pagination_control_frame.pack(pady=(5, 5)) 
    
    notebook = ttk.Notebook(main_container); notebook.pack(expand=True, fill='both')
    
    bottom_control_frame = tk.Frame(main_container, pady=10)
    bottom_control_frame.pack(fill='x', anchor='s')
    download_button = ttk.Button(
        bottom_control_frame, 
        text="📥 Download Excel", 
        command=lambda: export_to_excel(result_window.vehicle_data_list, date_str, lokasi_cabang)
    )
    download_button.pack()
    
    result_window.master_vehicle_list = parsed_data[:]
    result_window.vehicle_data_list = parsed_data[:]
    result_window.current_page = 0
    
    # --- FUNGSI UTAMA UNTUK MERENDER ULANG TAMPILAN ---
    def refresh_display(page_to_show=0):
        # 1. Hapus semua tab yang ada
        for tab in notebook.tabs():
            notebook.forget(tab)

        # 2. Putuskan apakah tab rangkuman perlu ditampilkan
        is_searching = bool(search_entry.get().strip())
        if not is_searching:
            create_summary_tab(notebook, result_window.vehicle_data_list)
        
        # 3. Buat tab kendaraan untuk halaman yang diminta
        start_index = page_to_show * VEHICLES_PER_PAGE
        end_index = start_index + VEHICLES_PER_PAGE
        vehicles_to_display = result_window.vehicle_data_list[start_index:end_index]
        for v_data in vehicles_to_display:
            create_vehicle_tab(notebook, v_data)
        
        # 4. Update kontrol pagination
        for widget in pagination_control_frame.winfo_children():
            widget.destroy()
            
        total_pages = (len(result_window.vehicle_data_list) + VEHICLES_PER_PAGE - 1) // VEHICLES_PER_PAGE
        if total_pages > 1:
            left_button = ttk.Button(pagination_control_frame, text="<", command=lambda: move_page(-1))
            left_button.pack(side=tk.LEFT, padx=(0, 5))
            if page_to_show == 0: left_button.config(state=tk.DISABLED)
            
            ttk.Label(pagination_control_frame, text=f"Halaman {page_to_show + 1}/{total_pages}", font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
            
            right_button = ttk.Button(pagination_control_frame, text=">", command=lambda: move_page(1))
            right_button.pack(side=tk.LEFT, padx=(5, 0))
            if page_to_show == total_pages - 1: right_button.config(state=tk.DISABLED)

        # 5. Pilih tab yang aktif
        if not is_searching:
            notebook.select(0)
        elif notebook.tabs():
            notebook.select(0)

    # --- Fungsi yang dipanggil oleh UI ---
    # --- Fungsi yang dipanggil oleh UI ---
    def perform_global_search(event=None):
        search_term = search_entry.get().strip().lower()
        
        if not search_term:
            # Jika tidak ada kata kunci, tampilkan semua data master
            result_window.vehicle_data_list = result_window.master_vehicle_list[:]
        else:
            filtered_list = []
            for vehicle in result_window.master_vehicle_list:
                vehicle_name = vehicle.get('vehicleName', '').lower()
                
                # 1. Cek apakah nama kendaraan cocok
                if search_term in vehicle_name:
                    # Jika cocok, tambahkan seluruh data kendaraan (termasuk semua stop)
                    filtered_list.append(vehicle)
                else:
                    # 2. Jika nama kendaraan tidak cocok, cek setiap stop
                    matching_stops = []
                    for stop in vehicle.get('stopDetails', []):
                        so_numbers = stop.get('soNumbers', '').lower()
                        customer_name = stop.get('visitName', '').lower() # Ini adalah 'Outlet'
                        
                        # Cek apakah SO atau Nama Outlet (Customer) cocok
                        if search_term in so_numbers or search_term in customer_name:
                            matching_stops.append(stop)
                    
                    # Jika ada stop yang cocok, buat data kendaraan baru
                    # hanya dengan stop yang cocok tersebut
                    if matching_stops:
                        filtered_vehicle_data = vehicle.copy()
                        filtered_vehicle_data['stopDetails'] = matching_stops
                        filtered_list.append(filtered_vehicle_data)
                        
            result_window.vehicle_data_list = filtered_list
            
        result_window.current_page = 0
        refresh_display(page_to_show=0)

    def reset_global_search():
        search_entry.delete(0, tk.END)
        perform_global_search()

    def move_page(delta):
        new_page = result_window.current_page + delta
        result_window.current_page = new_page
        refresh_display(page_to_show=new_page)

    search_button = ttk.Button(search_frame, text="Cari", command=perform_global_search)
    search_button.pack(side=tk.LEFT, padx=(5, 0))
    reset_button = ttk.Button(search_frame, text="Reset", command=reset_global_search)
    reset_button.pack(side=tk.LEFT, padx=(5, 0))
    search_entry.bind("<Return>", perform_global_search)
    
    # Simpan fungsi move_page agar bisa diakses
    result_window.move_page = move_page
    
    # Tampilan Awal
    refresh_display()

# (Sisa kode di bawah ini tidak ada perubahan)
# =============================================================================
def _parse_delivery_data(routing_results):
    parsed_data = []
    for route_item in routing_results:
        for route in route_item.get('result', {}).get('routing', []):
            vehicle_name = route.get('vehicleName', 'N/A')
            trips = route.get('trips', [])
            non_hub_trips = [trip for trip in trips if not trip.get('isHub')]
            num_trips = len(non_hub_trips)
            stop_details = []
            for trip in trips:
                time_window = trip.get('timeWindow', {})
                outlet_name = ""
                so_numbers = ""
                visit_name_raw = trip.get('visitName', '')
                
                is_hub_trip = trip.get('isHub', False)
                if is_hub_trip:
                    outlet_name = "HUB"
                    if not trip.get('eta') and not trip.get('etd'):
                        continue
                else:
                    if not visit_name_raw or not time_window.get('startTime') or not trip.get('eta'):
                        continue
                    outlet_name = extract_outlet_name(visit_name_raw)
                    
                    match = re.search(r'S[SO]\d', visit_name_raw)
                    if match:
                        so_start_index = match.start()
                        so_substring = visit_name_raw[so_start_index:]
                        so_list = [so.strip() for so in so_substring.split(',')]
                        so_numbers = ', '.join(so_list)

                stop_details.append({
                    "order": trip.get('order', 999),
                    "visitName": outlet_name,
                    "soNumbers": so_numbers, 
                    "startTime": time_window.get('startTime', 'N/A'),
                    "endTime": time_window.get('endTime', 'N/A'),
                    "eta": trip.get('eta', 'N/A'),
                    "etd": trip.get('etd', 'N/A'),
                })
            if stop_details:
                parsed_data.append({
                    "vehicleName": vehicle_name,
                    "numTrips": num_trips,
                    "stopDetails": stop_details
                })
    return parsed_data

def _handle_api_request_and_parse_data(app_instance, date_obj, hub_id):
    secret = load_secret()
    constants = load_constants()
    base_url = constants.get('base_url')
    token = secret.get('token')
    if not base_url or not token:
        show_error_message("Error API", ERROR_MESSAGES.get("API_TOKEN_MISSING", "Token API hilang."))
        return None, None
    day_of_week = date_obj.weekday()
    if day_of_week == 0:
        target_date_obj = date_obj - timedelta(days=2)
    else:
        target_date_obj = date_obj - timedelta(days=1)
    mileapp_date_format = target_date_obj.strftime('%Y-%m-%d')
    date_str = date_obj.strftime('%d-%m-%Y')
    url = f"{base_url}/results"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    params = {'dateFrom': mileapp_date_format, 'dateTo': mileapp_date_format, 'limit': 100, 'hubId': hub_id}
    response = requests.get(url, headers=headers, params=params, timeout=30)
    response.raise_for_status()
    data = response.json()
    app_instance.update_status("Memfilter data...")
    routing_results = [
        item for item in data.get('data', {}).get('data', [])
        if item.get("dispatchStatus") == "done"
    ]
    if not routing_results:
        app_instance.display_error("Data Tidak Ditemukan", ERROR_MESSAGES.get("DATA_NOT_FOUND", "Data tidak ditemukan."))
        return None, date_str
    app_instance.update_status("Mengekstrak dan memformat data estimasi...")
    parsed_data = _parse_delivery_data(routing_results)
    def get_hub_departure_time(vehicle):
        try:
            hub_stops = [s for s in vehicle.get('stopDetails', []) if s.get('visitName') == 'HUB']
            if not hub_stops: return datetime.max
            first_hub_stop = min(hub_stops, key=lambda s: s.get('order', 999))
            if first_hub_stop and first_hub_stop.get('etd') != 'N/A':
                return datetime.strptime(first_hub_stop['etd'], "%H:%M:%S")
        except (ValueError, KeyError):
            pass
        return datetime.max
    parsed_data.sort(key=lambda v: (get_hub_departure_time(v), v.get('vehicleName', '')))
    return parsed_data, date_str

def process_data(date_input, app_instance):
    date_str_input = date_input.get('dmy') if isinstance(date_input, dict) else (date_input if isinstance(date_input, str) else None)
    if not date_str_input:
        app_instance.display_error("Kesalahan Input", "Input tanggal tidak valid. Proses dibatalkan.")
        app_instance.after(1000, app_instance.destroy) 
        return
    try:
        date_obj = datetime.strptime(date_str_input, '%d-%m-%Y') 
        if date_obj.weekday() == 6:
            app_instance.display_error("Data Tidak Ditemukan", ERROR_MESSAGES.get("DATA_NOT_FOUND", "Data tidak ditemukan."))
            app_instance.after(1000, app_instance.destroy) 
            return
            
        config = load_config()
        constants = load_constants()
        lokasi_code = config.get('lokasi', 'Unknown')
        
        location_id_map = constants.get('location_id', {})
        reversed_location_map = {v: k for k, v in location_id_map.items()}
        lokasi_cabang = reversed_location_map.get(lokasi_code, lokasi_code)
        
        hub_id = get_hub_id() 
        if not hub_id:
            app_instance.after(1000, app_instance.destroy) 
            return
            
        parsed_data, date_str = _handle_api_request_and_parse_data(app_instance, date_obj, hub_id)
        
        if parsed_data:
            app_instance.after(0, lambda: display_result_gui(app_instance, parsed_data, date_str, lokasi_cabang))
        elif date_str:
            app_instance.display_error("Data Kosong", "Tidak ada kendaraan yang lolos filter atau tidak ada data estimasi waktu yang lengkap.")
            app_instance.after(1000, app_instance.destroy)
            
    except requests.exceptions.RequestException as e:
        handle_requests_error(e)
        app_instance.after(1000, app_instance.destroy)
    except Exception as e:
        error_msg = ERROR_MESSAGES.get("UNKNOWN_ERROR", "Kesalahan tak terduga: {error_detail}").format(error_detail=str(e))
        app_instance.display_error("Kesalahan Tak Terduga", error_msg)
        app_instance.after(1000, app_instance.destroy) 

def main():
    create_date_picker_window("Estimasi Delivery", process_data)