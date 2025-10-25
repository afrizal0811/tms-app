from datetime import datetime
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import pandas as pd
import re
import math

from utils.function import (
    get_save_path,
    load_config,
    load_constants,
    load_master_data,
    open_file_externally,
    show_error_message,
    show_info_message
)
from utils.messages import ERROR_MESSAGES, INFO_MESSAGES

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def parse_latlon(latlon_str):
    """Mengurai string 'lat,lon' menjadi float."""
    if pd.isna(latlon_str) or latlon_str in ['', '-']:
        return None, None
    try:
        parts = str(latlon_str).split(',')
        if len(parts) != 2:
            return None, None
        lat = float(parts[0].strip())
        lon = float(parts[1].strip())
        return lat, lon
    except (ValueError, TypeError):
        return None, None

def calculate_distance(latlon1_str, latlon2_str):
    """Menghitung jarak Haversine antara dua string lat/lon dalam meter."""
    lat1, lon1 = parse_latlon(latlon1_str)
    lat2, lon2 = parse_latlon(latlon2_str)

    # Jika salah satu koordinat tidak valid, kembalikan string kosong
    if lat1 is None or lat2 is None:
        return ""

    R = 6371000  # Radius Bumi dalam meter

    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)

    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad

    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    distance = R * c
    return int(distance) # Kembalikan jarak dalam meter (bulat)

def apply_styles_and_formatting(writer):
    workbook = writer.book
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cols_to_center = [
        'Open Time', 'Close Time', 'ETA', 'ETD', 'Actual Arrival',
        'Actual Departure', 'Visit Time', 'Actual Visit Time',
        'Customer ID', 'RO Sequence', 'Real Sequence', 'Temperature',
        'Total Visit', 'Total Delivered', 'Status Delivery', 'Is Same Sequence'
    ]

    for sheet_name in workbook.sheetnames:
        worksheet = writer.sheets[sheet_name]
        header_map = {cell.value: cell.column for cell in worksheet[1]}
        for col_name, col_idx in header_map.items():
            col_letter = get_column_letter(col_idx)
            align = center_align if col_name in cols_to_center else left_align
            for cell in worksheet[col_letter]:
                cell.alignment = align
                if col_name == ' ':
                    cell.fill = red_fill
        if ' ' in header_map:
            sep_col_idx = header_map[' ']
            worksheet.cell(row=1, column=sep_col_idx).value = ""
        for column_cells in worksheet.columns:
            try:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 50)
            except ValueError:
                pass

def convert_datetime_column(df, column_name, target_format='%H:%M'):
    def convert(val):
        if pd.isna(val) or val == '': return ''
        try:
            if isinstance(val, datetime):
                dt = val
            elif 'T' in str(val):
                dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
            else:
                dt = pd.to_datetime(val)
            return dt.strftime(target_format)
        except Exception:
            return val
    df[column_name] = df[column_name].apply(convert)
    return df

def calculate_actual_visit(start, end):
    if start == '' or end == '' or pd.isna(start) or pd.isna(end): 
        return ""
    try:
        t1 = datetime.strptime(str(start), "%H:%M")
        t2 = datetime.strptime(str(end), "%H:%M")
        delta = (t2 - t1).total_seconds()
        if delta < 0: delta += 86400
        return int(delta // 60)
    except (ValueError, TypeError):
        return ""

# =============================================================================
# DATAFRAME PROCESSING
# =============================================================================

def process_total_delivered(df, master_driver_df):
    # 1. Buat master summary, dikunci berdasarkan Nama Driver
    master_summary = master_driver_df[['Driver', 'Plat']].drop_duplicates().rename(columns={'Plat': 'License Plat'})
    
    # Cek jika kolom yang dibutuhkan ada
    if 'assignee' not in df.columns:
        # Jika tidak ada data, kembalikan daftar master kosong
        final_df = master_summary.assign(**{'Total Visit': pd.NA, 'Total Delivered': pd.NA})
        final_df[['Total Visit','Total Delivered']] = final_df[['Total Visit','Total Delivered']].astype('Int64')
        return final_df[['License Plat','Driver','Total Visit','Total Delivered']].sort_values('Driver')

    # Buat salinan df untuk diproses
    df_proc = df.copy()

    # 2. Map email 'assignee' ke 'Driver' (Nama Driver)
    email_to_name = dict(zip(master_driver_df['Email'].str.lower(), master_driver_df['Driver']))
    df_proc['Driver'] = df_proc['assignee'].str.lower().map(email_to_name)
    
    # Hapus baris yang tidak memiliki mapping driver (bukan bagian dari master)
    df_proc.dropna(subset=['Driver'], inplace=True)
    
    if df_proc.empty:
        # Jika setelah difilter jadi kosong, kembalikan master kosong
        final_df = master_summary.assign(**{'Total Visit': pd.NA, 'Total Delivered': pd.NA})
        final_df[['Total Visit','Total Delivered']] = final_df[['Total Visit','Total Delivered']].astype('Int64')
        return final_df[['License Plat','Driver','Total Visit','Total Delivered']].sort_values('Driver')


    # 3. Hitung Total Visit per Nama Driver
    visit_counts = df_proc['Driver'].value_counts(dropna=True).reset_index()
    visit_counts.columns = ['Driver', 'Total Visit']

    # 4. Hitung task yang "Tidak Terkirim" (PENDING, BATAL, TERIMA SEBAGIAN) per Nama Driver
    
    # Cari kolom status yang benar ('Status Delivery' atau 'label')
    if 'Status Delivery' in df_proc.columns:
        status_col_name = 'Status Delivery'
    elif 'label' in df_proc.columns:
        status_col_name = 'label'
    else:
        status_col_name = None

    if status_col_name:
        # Buat list dari status (mengatasi jika ada status ganda seperti "PENDING; SUKSES")
        df_proc['status_list'] = df_proc[status_col_name].fillna('').astype(str).str.upper().str.split(';')
        
        def is_not_delivered(status_list):
            if not isinstance(status_list, list):
                return False
            # Cek apakah 'PENDING', 'BATAL', atau 'TERIMA SEBAGIAN' ada di dalam list status
            statuses = {s.strip() for s in status_list}
            return 'PENDING' in statuses or 'BATAL' in statuses or 'TERIMA SEBAGIAN' in statuses
        
        df_proc['is_not_delivered'] = df_proc['status_list'].apply(is_not_delivered)
        
        # Hitung jumlah task yg 'is_not_delivered' per Nama Driver
        not_delivered_counts = df_proc[df_proc['is_not_delivered'] == True].groupby('Driver').size().reset_index(name='Total Not Delivered')
    else:
        # Jika tidak ada kolom label, buat dataframe kosong
        not_delivered_counts = pd.DataFrame(columns=['Driver', 'Total Not Delivered'])

    # 5. Gabungkan hasil hitungan
    # Mulai dengan visit_counts, gabungkan not_delivered_counts
    # Kode yang BENAR
    final_counts = visit_counts.merge(not_delivered_counts, on='Driver', how='left')
    # Isi driver yang tidak punya status PENDING/BATAL/TERIMA SEBAGIAN dengan 0
    final_counts['Total Not Delivered'] = final_counts['Total Not Delivered'].fillna(0)
    
    # 6. Hitung Total Delivered sesuai logika baru
    final_counts['Total Delivered'] = final_counts['Total Visit'] - final_counts['Total Not Delivered']

    # 7. Gabungkan dengan master_summary
    # Gunakan 'left' join dari master_summary untuk menyertakan semua driver dari master
    final_df = master_summary.merge(final_counts, on='Driver', how='left')

    # 8. Finalisasi
    # Buang kolom bantu
    final_df = final_df.drop(columns=['Total Not Delivered'], errors='ignore')
    # Ubah tipe data kolom agar NA tertulis rapi di Excel
    final_df[['Total Visit','Total Delivered']] = final_df[['Total Visit','Total Delivered']].astype('Int64')
    
    # Kembalikan dataframe dengan urutan kolom yang benar
    return final_df[['License Plat','Driver','Total Visit','Total Delivered']].sort_values('Driver')

def process_ro_vs_real(df, master_driver_df):
    df_proc = df.copy()
    
    # 1. Mappings (License Plat & Driver)
    email_to_name = dict(zip(master_driver_df['Email'].str.lower(), master_driver_df['Driver']))
    email_to_plat = dict(zip(master_driver_df['Email'].str.lower(), master_driver_df['Plat']))
    
    df_proc['assignee_email'] = df_proc['assignee'].str.lower()
    
    # + Driver (Lookup)
    df_proc['Driver'] = df_proc['assignee_email'].map(email_to_name).fillna(df_proc['assignee'])
    
    # + License Plat (Lookup dengan fallback)
    df_proc['License Plat'] = df_proc['assignee_email'].map(email_to_plat)
    if 'assignedVehicle' in df_proc.columns:
        df_proc['License Plat'] = df_proc['License Plat'].fillna(df_proc['assignedVehicle'])

    # 2. Convert Time Columns
    # + Actual Arrival (ATURAN 2)
    # Tentukan sumber 'Actual Arrival' berdasarkan 'flow'
    if 'flow' in df_proc.columns and 'page1DoneTime' in df_proc.columns:
        df_proc['Actual Arrival_Source'] = df_proc.apply(
            lambda x: x['page1DoneTime'] if 'Pending GR' in str(x['flow']) else x['Klik Jika Sudah Sampai'],
            axis=1
        )
    else:
        # Fallback jika 'flow' or 'page1DoneTime' tidak ada
        df_proc['Actual Arrival_Source'] = df_proc['Klik Jika Sudah Sampai']

    # Konversi kolom waktu
    for col in ['Actual Arrival_Source', 'doneTime', 'page1DoneTime']:
        if col in df_proc.columns:
            convert_datetime_column(df_proc, col)

    # Rename kolom sumber ke 'Actual Arrival'
    df_proc.rename(columns={'Actual Arrival_Source': 'Actual Arrival'}, inplace=True)

    # 3. Calculate Visit Time
    # + Actual Visit Time
    df_proc['Actual Visit Time'] = df_proc.apply(
        lambda r: calculate_actual_visit(r.get('Actual Arrival',''), r.get('doneTime','')), axis=1
    )

    # 4. Handle Sequences
    # + Real Sequence (Rank by new 'Actual Arrival')
    df_proc['Actual Arrival_parsed'] = pd.to_datetime(df_proc['Actual Arrival'], format='%H:%M', errors='coerce')
    df_proc['Real Sequence'] = df_proc.groupby('Driver')['Actual Arrival_parsed'].rank(method='dense').astype('Int64')

    # ==========================================================
    # 5. Handle Status Delivery (PERBAIKAN: Selalu dari 'label')
    # ==========================================================
    if 'label' in df_proc.columns:
        df_proc['Status Delivery'] = df_proc['label']
    else:
        df_proc['Status Delivery'] = '' # Kolom 'label' tidak ada
    # ==================== AKHIR PERBAIKAN ====================


    # 6. Rename kolom-kolom
    df_proc.rename(columns={
        'title': 'Customer',
        'doneTime': 'Actual Departure',
        'routePlannedOrder': 'RO Sequence'
    }, inplace=True)
    
    # Hapus kolom duplikat (penting jika file asli punya 'Status Delivery' DAN 'label')
    if df_proc.columns.duplicated().any():
        df_proc = df_proc.loc[:, ~df_proc.columns.duplicated()]

    # 7. + Is Same Sequence (ATURAN 3)
    ro_num = pd.to_numeric(df_proc['RO Sequence'], errors='coerce')
    real_num = pd.to_numeric(df_proc['Real Sequence'], errors='coerce')
    
    comparison = (ro_num == real_num)
    df_proc['Is Same Sequence'] = comparison.map({True: 'SAMA', False: 'TIDAK SAMA'})
    
    df_proc.loc[df_proc['RO Sequence'] == '-', 'Is Same Sequence'] = 'TIDAK SAMA'
    
    # Mengatasi FutureWarning
    df_proc['Is Same Sequence'] = df_proc['Is Same Sequence'].fillna('TIDAK SAMA')

    # 8. Tentukan Kolom Final
    cols = [
        'flow', 'License Plat', 'Driver', 'Customer', 'Status Delivery', 
        'Open Time', 'Close Time', 'Actual Arrival', 'Actual Departure', 
        'Visit Time', 'Actual Visit Time', 'RO Sequence', 'Real Sequence', 'Is Same Sequence'
    ]
    
    # Pastikan semua kolom yang diminta ada
    for col in cols:
        if col not in df_proc.columns:
            df_proc[col] = '' 

    # 9. Buat DataFrame Final dan Urutkan
    df_final = df_proc[cols].sort_values(['Driver', 'Real Sequence'])

    # 10. Tambahkan Baris Kosong Antar Grup
    parts = []
    df_final['Driver'] = df_final['Driver'].astype(str).str.strip()

    if df_final.empty:
        return pd.DataFrame(columns=cols) 

    for _, g in df_final.dropna(subset=['Driver']).groupby('Driver'):
        parts.append(g)
        row_dummy = pd.DataFrame({col: [''] for col in df_final.columns}).astype(object)
        row_dummy.iloc[0] = None
        parts.append(row_dummy)

    if parts and parts[-1].isna().all(axis=1).all():
        parts = parts[:-1]

    for i in range(len(parts)):
        parts[i] = parts[i].where(pd.notnull(parts[i]), None)

    safe_parts = [
        df if not df.isna().all(axis=None) else pd.DataFrame([{col: "" for col in df.columns}])
        for df in parts
    ]
    
    return pd.concat(safe_parts, ignore_index=True)

def process_pending_so(df, master_driver_df):
    df_proc = df.copy()
    
    # Mapping untuk License Plat dan Driver Name
    email_to_name = dict(zip(master_driver_df['Email'].str.lower(), master_driver_df['Driver']))
    email_to_plat = dict(zip(master_driver_df['Email'].str.lower(), master_driver_df['Plat']))

    df_proc['assignee_lower'] = df_proc['assignee'].str.lower()
    
    # + Driver: Menggunakan NAMA DRIVER dari master (sesuai klarifikasi sebelumnya)
    df_proc['Driver'] = df_proc['assignee_lower'].map(email_to_name).fillna(df_proc['assignee'])
    
    # + License Plat: Menggunakan PLAT dari master
    df_proc['License Plat'] = df_proc['assignee_lower'].map(email_to_plat)
    # Jika tidak ada di master, gunakan 'assignedVehicle' dari file
    if 'assignedVehicle' in df_proc.columns:
        df_proc['License Plat'] = df_proc['License Plat'].fillna(df_proc['assignedVehicle'])
    
    # --- START PERBAIKAN STATUS ---
    # Cari kolom status utama ('Status Delivery' atau 'label')
    if 'Status Delivery' in df_proc.columns:
        status_col = df_proc['Status Delivery']
    elif 'label' in df_proc.columns:
        status_col = df_proc['label']
    else:
        status_col = pd.Series(index=df_proc.index, dtype='str') 

    df_proc['status_delivery_list'] = status_col.fillna('').astype(str).str.upper().str.split(';')
    df_proc['status_delivery_list'] = df_proc['status_delivery_list'].apply(lambda x: list(dict.fromkeys([s.strip() for s in x if s.strip()])))
    
    # --- END PERBAIKAN STATUS ---

    # Filter semua status relevan (termasuk dari status_delivery_list)
    status_to_filter = ['BATAL', 'TERIMA SEBAGIAN', 'PENDING', 'PENDING GR']
    
    # Cek juga dari kolom 'Status GR' jika ada
    if 'Status GR' in df_proc.columns:
        # Tambahkan status dari 'Status GR' ke 'status_delivery_list'
        status_gr_list = df_proc['Status GR'].fillna('').astype(str).str.upper().str.split(';')
        status_gr_list = status_gr_list.apply(lambda x: list(dict.fromkeys([s.strip() for s in x if s.strip()])))
        
        # Gabungkan list status
        df_proc['status_delivery_list'] = df_proc['status_delivery_list'] + status_gr_list
        df_proc['status_delivery_list'] = df_proc['status_delivery_list'].apply(lambda x: list(dict.fromkeys(x))) # Re-uniquify
        
        # Filter juga baris yang 'Status GR' nya PENDING GR
        df_filtered = df_proc[
            df_proc['status_delivery_list'].apply(lambda lst: any(s in status_to_filter for s in lst))
        ].copy()
    else:
        # Filter seperti biasa jika 'Status GR' tidak ada
        df_filtered = df_proc[
            df_proc['status_delivery_list'].apply(lambda lst: any(s in status_to_filter for s in lst))
        ].copy()

    if df_filtered.empty:
        return None

    # + Actual Arrival, Actual Departure, ETA, ETD (Konversi Waktu)
    # Menggunakan 'doneTime' sesuai kolom yang ada di file
    for col in ['Klik Jika Sudah Sampai', 'doneTime', 'eta', 'etd']:
        if col in df_filtered.columns:
            convert_datetime_column(df_filtered, col)
        
    # + Actual Visit Time
    df_filtered['Actual Visit Time'] = df_filtered.apply(
        lambda r: calculate_actual_visit(r.get('Klik Jika Sudah Sampai',''), r.get('doneTime','')), axis=1
    )

    # + Customer ID (menggunakan regex 'C0')
    def get_customer_id(title):
        match = re.search(r'(C0\d+)', str(title))
        return match.group(1) if match else ''
    df_filtered['Customer ID'] = df_filtered['title'].apply(get_customer_id)

    # + Temperature
    df_filtered['Temperature'] = df_filtered['Driver'].astype(str).str.split(' ').str[0].str.replace("'", "")

    # + Reason (Menggunakan logika OR yang lebih robust)
    def get_reason(row):
        # Hanya ambil alasan jika ada status relevan
        if any(s in status_to_filter for s in row['status_delivery_list']):
            return row.get('Alasan') or row.get('Alasan Tidak Bisa Dikunjungi') or row.get('Alasan Batal') or ''
        return ''
    df_filtered['Reason'] = df_filtered.apply(get_reason, axis=1)

    # + Real Sequence (BARU)
    # Parse 'Actual Arrival' untuk perbandingan
    df_filtered['Actual Arrival_parsed'] = pd.to_datetime(df_filtered['Klik Jika Sudah Sampai'], format='%H:%M', errors='coerce')
    df_filtered['Real Sequence'] = df_filtered.groupby('Driver')['Actual Arrival_parsed'].rank(method='dense').astype('Int64')

    # --- Pemrosesan kolom Status Faktur ---
    def assign_status_columns(row):
        statuses = row['status_delivery_list']
        title = row['title']
        
        # + Faktur Batal/ Tolakan SO
        faktur_batal = title if "BATAL" in statuses else ''
        
        # + Terkirim Sebagian
        terkirim_sebagian = title if "TERIMA SEBAGIAN" in statuses else ''
        
        # + Pending (Hanya PENDING, tidak PENDING GR)
        pending = title if "PENDING" in statuses else ''
        
        # + Pending GR (Bisa dari 'Status GR' atau 'Status Delivery'/'label')
        pending_gr = title if "PENDING GR" in statuses else ''

        return faktur_batal, terkirim_sebagian, pending, pending_gr

    (df_filtered['Faktur Batal/ Tolakan SO'],
     df_filtered['Terkirim Sebagian'],
     df_filtered['Pending'],
     df_filtered['Pending GR']) = zip(*df_filtered.apply(assign_status_columns, axis=1))
    
    # --- Daftar kolom keluaran ---
    # Menyesuaikan urutan kolom sesuai permintaan
    cols = [
        'License Plat', 'Driver',
        'Faktur Batal/ Tolakan SO', 'Terkirim Sebagian', 'Pending', 'Pending GR',
        'Reason', 'Open Time', 'Close Time', 'eta', 'etd',
        'Klik Jika Sudah Sampai', 'doneTime', 'Visit Time', 'Actual Visit Time',
        'Customer ID', 'routePlannedOrder', 'Real Sequence', 'Temperature'
    ]
    
    # Rename kolom agar sesuai output
    df_final = df_filtered[cols].rename(columns={
        'eta': 'ETA', 'etd': 'ETD',
        'Klik Jika Sudah Sampai': 'Actual Arrival',
        'doneTime': 'Actual Departure',
        'routePlannedOrder': 'RO Sequence'
    })
    
    # Tambahkan kolom separator ' '
    reason_loc = df_final.columns.get_loc('Reason')
    if ' ' not in df_final.columns:
        df_final.insert(reason_loc + 1, ' ', '')
        
    # Mengelompokkan (mengurutkan) berdasarkan Driver
    return df_final.sort_values(['Driver', 'Real Sequence'])

def process_update_longlat(df):
    # Tentukan kolom input dan output
    required_cols = ['title', 'Klik Lokasi Client', 'Longlat']
    output_columns = ["Customer ID", "Customer Name", "Location ID", "New Longlat", "Beda Jarak (m)"]

    # Jika kolom input penting tidak ada, kembalikan DataFrame kosong
    if not all(col in df.columns for col in required_cols):
        return pd.DataFrame(columns=output_columns)
    
    data = []
    
    for _, row in df.iterrows():
        
        # 1. Ambil nilai mentah dari 'Klik Lokasi Client'
        new_longlat_raw = row.get('Klik Lokasi Client')

        # 2. Cek apakah nilainya 'missing' (pd.NA, np.nan, None)
        if pd.isna(new_longlat_raw):
            continue
            
        # 3. Jika tidak missing, SEKARANG baru ubah ke string dan strip
        new_longlat = str(new_longlat_raw).strip()
        
        # 4. Cek untuk string kosong atau '-'
        if new_longlat in ['', '-']:
            continue
        
        # Kode sisa aman karena 'new_longlat' dijamin string yg valid
        old_longlat = str(row.get('Longlat', '')).strip()
        title_str = str(row['title'])
        
        # Parse title
        parts = [p.strip() for p in title_str.split('-')]
        
        # Ambil Kode C0...
        match = re.search(r'(C0\d+)', title_str)
        customer_id_code = match.group(1) if match else ''
        
        # Ambil bagian kiri strip pertama
        customer_name_str = parts[0] if parts else ''
        
        # Ambil bagian kanan strip terakhir
        location_id_val = parts[-1] if len(parts) > 1 else ''

        # Beda Jarak (m)
        beda_jarak = calculate_distance(old_longlat, new_longlat)

        data.append({
            "Customer ID": customer_id_code,
            "Customer Name": customer_name_str,
            "Location ID": location_id_val,
            "New Longlat": new_longlat,
            "Beda Jarak (m)": beda_jarak
        })
    
    # Jika data kosong (tidak ada update longlat), kembalikan DataFrame kosong
    if not data:
        return pd.DataFrame(columns=output_columns)

    # ==========================================================
    # ===== PERUBAHAN (Buat DF, Konversi Tipe, dan Sortir) =====
    # ==========================================================
    
    # 1. Buat DataFrame dari list
    df_final = pd.DataFrame(data, columns=output_columns)
    
    # 2. Ubah 'Beda Jarak (m)' ke numerik agar bisa disortir (string kosong akan jadi NaN)
    df_final["Beda Jarak (m)"] = pd.to_numeric(df_final["Beda Jarak (m)"], errors='coerce')
    
    # 3. Sortir berdasarkan 'Beda Jarak (m)' secara ascending
    df_final.sort_values(by="Beda Jarak (m)", ascending=True, inplace=True)
    
    # 4. Kembalikan DataFrame yang sudah disortir
    return df_final

def get_created_date(file_path):
    try:
        df_main = pd.read_excel(file_path, sheet_name="Main")
        if "startTime" not in df_main.columns:
            return datetime.now()

        start_val = df_main["startTime"].dropna().iloc[0]

        if isinstance(start_val, str):
            try:
                dt = datetime.strptime(start_val.strip(), "%Y-%m-%d %H:%M")
            except ValueError:
                dt = pd.to_datetime(start_val, errors="coerce")
        elif isinstance(start_val, datetime):
            dt = start_val
        else:
            dt = pd.to_datetime(str(start_val), errors="coerce")

        if pd.isna(dt):
            return datetime.now()

        return dt
    except Exception:
        return datetime.now()


# =============================================================================
# MAIN
# =============================================================================

def main():
    config = load_config()
    constants = load_constants()
    if not config or "lokasi" not in config:
        show_error_message("Dibatalkan", ERROR_MESSAGES["LOCATION_CODE_MISSING"]); return
    lokasi_code = config["lokasi"]
    show_info_message("Upload File Task", INFO_MESSAGES["SELECT_FILE"].format(text="export task"))
    input_file = filedialog.askopenfilename(title="Pilih File Excel yang Akan Diproses", filetypes=[("Excel Files","*.xlsx *.xls")])
    if not input_file:
        show_info_message("Dibatalkan", INFO_MESSAGES["CANCELED_BY_USER"]); return
    df_original = pd.read_excel(input_file)
    required_columns = ['assignedVehicle','assignee','Alasan Tidak Bisa Dikunjungi','Alasan Batal','Open Time','Close Time','eta','etd','Klik Jika Sudah Sampai','doneTime','Visit Time','routePlannedOrder']
    if any(col not in df_original.columns for col in required_columns):
        show_error_message("Proses Gagal", ERROR_MESSAGES["INVALID_FILE"].format(details="Upload file Export Task dengan benar!")); return
    email_prefixes = df_original["assignee"].dropna().astype(str).str.extract(r'kendaraan\.([^.@]+)',expand=False).dropna().str.lower().unique()
    if not any(lokasi_code.lower() in prefix for prefix in email_prefixes):
        show_error_message("Proses Gagal", ERROR_MESSAGES["LOCATION_CODE_MISSING"]); return
    master_data = load_master_data(lokasi_code)
    if master_data is None:
        show_error_message("Proses Gagal", ERROR_MESSAGES["MASTER_DATA_MISSING"]); return

    master_df = master_data["df"]
    required_master_cols = {'Driver', 'Plat', 'Email'}
    if not required_master_cols.issubset(master_df.columns):
        show_error_message("Proses Gagal", "Kolom pada data master tidak lengkap."); return

    required_master_cols = {'Driver', 'Plat', 'Email'}
    if not required_master_cols.issubset(master_df.columns):
        show_error_message("Proses Gagal", "Kolom pada data master tidak lengkap."); return
    
    results_to_save = {
        'Total Delivered': process_total_delivered(df_original, master_df),
        'Hasil Pending SO': process_pending_so(df_original, master_df),
        'Hasil RO vs Real': process_ro_vs_real(df_original, master_df),
        'Update Longlat': process_update_longlat(df_original)
    }
    if results_to_save['Update Longlat'].empty:
        results_to_save['Update Longlat'] = pd.DataFrame([{"Customer ID":"Tidak Ada Update Longlat","Customer Name":"","Location ID":"","New Longlat":"", "Beda Jarak (m)":""}])
    location_id = constants.get('location_id', {})
    lokasi_name = next((n for n,c in location_id.items() if c == lokasi_code), lokasi_code)
    # input_filename = os.path.basename(input_file)
    created_date = get_created_date(input_file)
    date_str = created_date.strftime('%d.%m.%Y')
    file_basename = f"Delivery Summary - {date_str} - {lokasi_name}"
    save_file_path = get_save_path(file_basename)
    if not save_file_path: show_error_message("Proses Gagal", INFO_MESSAGES["CANCELED_BY_USER"]); return
    with pd.ExcelWriter(save_file_path, engine='openpyxl') as writer:
        for sheet in ['Total Delivered','Hasil Pending SO','Hasil RO vs Real','Update Longlat']:
            if sheet in results_to_save and results_to_save[sheet] is not None:
                results_to_save[sheet].to_excel(writer, sheet_name=sheet, index=False)
        apply_styles_and_formatting(writer)
    open_file_externally(save_file_path)

if __name__ == "__main__":
    main()