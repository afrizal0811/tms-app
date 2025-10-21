from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import numpy as np
import requests
import traceback
import sys
from utils.function import (
    get_save_path,
    load_config,
    load_constants,
    load_master_data,
    load_secret,
    load_type_map,
    open_file_externally,
    show_error_message,
    show_info_message
)
from utils.gui import create_date_picker_window
from utils.messages import ERROR_MESSAGES, INFO_MESSAGES
from utils.api_handler import handle_requests_error

project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.append(project_root)

def style_excel(file_path):
    workbook = load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    if "Truck Detail" in workbook.sheetnames:
        ws_detail = workbook["Truck Detail"]
        center_align = Alignment(horizontal='center', vertical='center')
        for row in ws_detail.iter_rows(min_row=2, min_col=3, max_col=8):
            for cell in row:
                cell.alignment = center_align
    workbook.save(file_path)

def process_routing_data(date_formats, gui_instance):
    try:
        selected_date_dmy = date_formats['dmy']
        selected_date_for_filename = selected_date_dmy.replace('-', '.')
        try:
            date_obj = datetime.strptime(selected_date_dmy, '%d-%m-%Y')
        except ValueError:
            show_error_message("Error Format Tanggal", "Format tanggal dari pemilih tidak valid.")
            return

        config = load_config()
        constants = load_constants()
        secrets = load_secret()
        lokasi_code = config.get('lokasi')
        
        # =======================================================
        # ▼▼▼ LOGIKA BARU UNTUK MENGGABUNGKAN PLCK + PLDM ▼▼▼
        # =======================================================
        codes_to_process = ['plck', 'pldm'] if lokasi_code == 'plck' else [lokasi_code]
        
        # Muat data master berdasarkan kode yang diproses
        # Asumsi: load_master_data() tanpa argumen akan memuat semua data
        full_master_data = load_master_data()
        if full_master_data is None: return

        # Filter master data sesuai dengan kode lokasi yang dibutuhkan
        master_data_df = full_master_data["df"]
        master_data_df = master_data_df[master_data_df['Email'].str.contains('|'.join(codes_to_process), na=False)]
        hub_ids = full_master_data["hub_ids"]
        # =======================================================
        
        if not config:
            show_error_message("Gagal", ERROR_MESSAGES["CONFIG_FILE_ERROR"])
            return
        if not constants:
            show_error_message("Gagal", ERROR_MESSAGES["CONSTANT_FILE_ERROR"])
            return
        if not secrets:
            show_error_message("Gagal", ERROR_MESSAGES["SECRET_FILE_ERROR"])
            return
        if master_data_df.empty:
            show_error_message("Gagal", ERROR_MESSAGES["MASTER_DATA_MISSING"])
            return

        master_data_map = dict(zip(master_data_df['Email'], master_data_df['Driver']))
        api_token = secrets.get('token')
        location_id = constants.get('location_id', {})
        
        # Tentukan nama lokasi untuk nama file (jika plck, tambahkan pldm)
        if lokasi_code == 'plck':
            lokasi_name = "PLCK & PLDM"
        else:
            lokasi_name = next((name for name, code in location_id.items() if code == lokasi_code), lokasi_code)

        if not api_token:
            show_error_message("Error Token API", ERROR_MESSAGES["API_TOKEN_MISSING"])
            return

        base_url = constants.get('base_url')
        api_url = f"{base_url}/results"
        headers = {'Authorization': f'Bearer {api_token}'}

        if date_obj.weekday() == 6:
            adjusted_date = date_obj
        else:
            adjusted_date = date_obj - timedelta(days=1)
            if adjusted_date.weekday() == 6:
                adjusted_date = date_obj - timedelta(days=2)

        date_str = adjusted_date.strftime('%Y-%m-%d')
        
        # =======================================================
        # ▼▼▼ LOGIKA PANGGILAN API BERULANG UNTUK SETIAP HUB ▼▼▼
        # =======================================================
        all_response_data = []
        hub_ids_to_fetch = [hub_ids.get(code) for code in codes_to_process if hub_ids.get(code)]

        if not hub_ids_to_fetch:
            show_error_message("Konfigurasi Salah", ERROR_MESSAGES["HUB_ID_MISSING"])
            return
            
        for hub_id in hub_ids_to_fetch:
            params = {
                'dateFrom': date_str, 'dateTo': date_str,
                'limit': 100, 'hubId': hub_id
            }
            try:
                response = requests.get(api_url, headers=headers, params=params, timeout=30)
                response.raise_for_status()
                data_check = response.json()
                if 'data' in data_check and 'data' in data_check['data'] and data_check['data']['data']:
                    all_response_data.extend(data_check['data']['data'])
            except requests.exceptions.RequestException as e:
                handle_requests_error(e)
                return
        
        if not all_response_data:
            show_error_message("Data Tidak Ditemukan", ERROR_MESSAGES["DATA_NOT_FOUND"])
            return
        # =======================================================

        processed_data = []
        first_tags_list = []
        processed_assignees_for_usage = set()

        routing_results = [
            item for item in all_response_data
            if item.get("dispatchStatus") == "done"
        ]
        
        for item in routing_results:
            if 'result' in item and 'routing' in item['result']:
                for route in item['result']['routing']:
                    assignee_email = route.get('assignee')
                    # Filter email tidak lagi diperlukan karena sudah difilter via hub_id dan master data
                    if not assignee_email:
                        continue

                    if assignee_email not in processed_assignees_for_usage:
                        vehicle_tags = route.get('vehicleTags', [])
                        if vehicle_tags:
                            first_tags_list.append(vehicle_tags[0])
                        processed_assignees_for_usage.add(assignee_email)

                    driver_name = master_data_map.get(assignee_email, assignee_email)
                    trips = route.get("trips", [])
                    def safe_float(value, default=0):
                        try:
                            return float(value)
                        except (TypeError, ValueError):
                            return default
    
                    if trips:
                        def is_hub_true(val):
                            if isinstance(val, bool): return val
                            if isinstance(val, str): return val.strip().lower() in ('true', '1', 'yes')
                            try: return int(val) == 1
                            except Exception: return False

                        non_hub_trips = [t for t in trips if not is_hub_true(t.get('isHub', False))]
                        total_weight = sum(safe_float(t.get("weight", 0)) for t in non_hub_trips)
                        total_volume = sum(safe_float(t.get("volume", 0)) for t in non_hub_trips)
                        total_distance = sum(safe_float(t.get("distance", 0)) for t in trips)
                        total_minutes = sum((t.get("travelTime", 0) + t.get("visitTime", 0) + t.get("waitingTime", 0)) for t in trips)
                        hours, minutes = divmod(total_minutes, 60)
                        ship_duration = f"'{hours}:{minutes:02d}"
                        vehicle_max_weight = route.get("vehicleMaxWeight", 1) or 1
                        vehicle_max_volume = route.get("vehicleMaxVolume", 1) or 1
                        weight_percentage = (total_weight / vehicle_max_weight) * 100 if vehicle_max_weight else None
                        volume_percentage = (total_volume / vehicle_max_volume) * 100 if vehicle_max_volume else None
                    else:
                        total_distance, ship_duration, weight_percentage, volume_percentage = None, None, None, None

                    processed_data.append({
                        'Assignee': driver_name,
                        'Vehicle Name': route.get('vehicleName'),
                        'Total Distance (m)': (round(total_distance) if total_distance is not None else None),
                        'Total Visits': None,
                        'Total Delivered': None,
                        'weight_numeric': weight_percentage,
                        'volume_numeric': volume_percentage,
                        'ship_duration': ship_duration
                    })

        if not processed_data:
            show_error_message("Data Tidak Ditemukan", "Tidak ada data 'done' yang ditemukan untuk diproses.")
            return

        df_api = pd.DataFrame(processed_data)
        df_api['Total Distance (m)'] = df_api['Total Distance (m)'].astype(object)

        def sum_or_none(series):
            values = series.dropna()
            return values.sum() if not values.empty else None

        agg_rules = {
            'Vehicle Name': lambda x: ', '.join(x.dropna().unique()),
            'Total Distance (m)': sum_or_none,
            'Total Visits': 'first',
            'Total Delivered': 'first',
            'weight_numeric': 'mean',
            'volume_numeric': 'mean',
            'ship_duration': 'first'
        }
        df_api_grouped = df_api.groupby('Assignee', as_index=False).agg(agg_rules)
        df_api_grouped['Weight Percentage'] = df_api_grouped['weight_numeric'].apply(lambda x: f"{x:.1f}%" if pd.notnull(x) else None)
        df_api_grouped['Volume Percentage'] = df_api_grouped['volume_numeric'].apply(lambda x: f"{x:.1f}%" if pd.notnull(x) else None)
        df_api_grouped.rename(columns={'ship_duration': 'Ship Duration'}, inplace=True)
        df_api_grouped = df_api_grouped.drop(columns=['weight_numeric', 'volume_numeric'])

        all_master_drivers = set(master_data_map.values())
        drivers_in_api = set(df_api_grouped['Assignee']) if not df_api_grouped.empty else set()
        missing_drivers = all_master_drivers - drivers_in_api
        df_missing = pd.DataFrame([{'Assignee': driver} for driver in missing_drivers])
        
        df_final = pd.concat([df_api_grouped, df_missing], ignore_index=True)

        # Blok pengurutan SEWA (tidak berubah)
        df_final['is_sewa'] = (
            df_final['Vehicle Name'].str.contains('SEWA', case=False, na=False) |
            df_final['Assignee'].str.contains('SEWA', case=False, na=False)
        ).astype(int)
        conditions = [
            df_final['Assignee'].str.contains('DRY', case=False, na=False),
            df_final['Assignee'].str.contains('FRZ', case=False, na=False)
        ]
        choices = [1, 2]
        df_final['sewa_category'] = np.select(conditions, choices, default=3)
        df_final = df_final.sort_values(
            by=['is_sewa', 'sewa_category', 'Assignee'],
            ascending=[True, True, True]
        ).reset_index(drop=True)
        df_final = df_final.drop(columns=['is_sewa', 'sewa_category'])

        column_order = ['Vehicle Name', 'Assignee', 'Weight Percentage', 'Volume Percentage', 'Total Distance (m)', 'Total Visits', 'Total Delivered', 'Ship Duration']
        df_final = df_final.reindex(columns=column_order)

        dry_dist_m = df_final[df_final['Assignee'].str.contains("DRY", na=False)]['Total Distance (m)'].sum(skipna=True)
        frz_dist_m = df_final[df_final['Assignee'].str.contains("FRZ", na=False)]['Total Distance (m)'].sum(skipna=True)
        df_summary = pd.DataFrame({'DRY': [round((dry_dist_m or 0) / 1000, 2)], 'FRZ': [round((frz_dist_m or 0) / 1000, 2)]})
        vehicle_types = constants.get("vehicle_types", [])
        usage_counts = {v_type: {'DRY': 0, 'FROZEN': 0} for v_type in vehicle_types}
        sorted_vehicle_types = sorted(vehicle_types, key=len, reverse=True)
        type_map = load_type_map().get("type", {}) if load_type_map() else {}

        for tag in first_tags_list:
            tag = type_map.get(tag, tag)
            category = None
            if "DRY" in tag:
                category = 'DRY'
            elif "FROZEN" in tag:
                category = 'FROZEN'

            if category:
                for v_type in sorted_vehicle_types:
                    if v_type in tag:
                        usage_counts[v_type][category] += 1
                        break

        usage_data_for_df = []
        for v_type, counts in usage_counts.items():
            dry_count = counts['DRY'] if counts['DRY'] > 0 else None
            frozen_count = counts['FROZEN'] if counts['FROZEN'] > 0 else None
            usage_data_for_df.append({'Tipe Kendaraan': v_type,'Jumlah (DRY)': dry_count,'Jumlah (FROZEN)': frozen_count})
        df_usage = pd.DataFrame(usage_data_for_df)

        file_basename = f"Routing Summary - {selected_date_for_filename} - {lokasi_name}"
        save_path = get_save_path(base_name=file_basename, extension=".xlsx")
        if not save_path:
            show_info_message("Dibatalkan", INFO_MESSAGES["CANCELED_BY_USER"])
            return

        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name='Truck Detail', index=False)
            df_summary.to_excel(writer, sheet_name='Total Distance Summary', index=False)
            df_usage.to_excel(writer, sheet_name='Truck Usage', index=False)

        style_excel(save_path)
        open_file_externally(save_path)

    except requests.exceptions.RequestException as e:
        handle_requests_error(e)
        return None
    except Exception as e:
        show_error_message("Error Tak Terduga", ERROR_MESSAGES["UNKNOWN_ERROR"].format(
        error_detail=f"{e}\n\n{traceback.format_exc()}"
    ))

def main():
    def process_wrapper(dates, gui_instance):
        def safe_close():
            if gui_instance and gui_instance.winfo_exists():
                gui_instance.destroy()
        try:
            process_routing_data(dates, gui_instance)
        finally:
            if gui_instance and gui_instance.winfo_exists():
                gui_instance.after(100, safe_close)
    create_date_picker_window(title="Routing Summary", process_callback=process_wrapper)

if __name__ == '__main__':
    main()