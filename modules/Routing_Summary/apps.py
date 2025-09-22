from datetime import datetime, timedelta
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import openpyxl
import pandas as pd
import re
import tkinter as tk
import traceback
from utils.function import (
    get_save_path,
    load_config,
    load_constants,
    load_master_data,
    load_type_map,
    open_file_externally,
    show_ask_message,
    show_error_message,
    show_info_message
)
from utils.messages import ASK_MESSAGES, ERROR_MESSAGES, INFO_MESSAGES

# ==============================================================================
# FUNGSI-FUNGSI UTAMA (HELPER FUNCTIONS)
# ==============================================================================

constants = load_constants()
def pilih_file_excel():
    root = tk.Tk()
    root.withdraw()
    show_info_message("Upload File", INFO_MESSAGES["SELECT_FILE"].format(text="export routing"))
    file_path = filedialog.askopenfilename(
        title="Upload File Routing",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    return file_path

def contains_capacity_constraint(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if any("capacity constraint" in str(cell).lower() for cell in row if cell):
                return True
    except Exception:
        return False
    return False

def buat_mapping_driver(lokasi_value):
    """Mapping email → driver name dari master.json berdasarkan lokasi"""
    master = load_master_data(lokasi_value)
    if not master or master["df"].empty:
        return {}
    df_driver = master["df"]
    return pd.Series(df_driver.Driver.values, index=df_driver.Email).to_dict()

def proses_truck_detail(workbook, source_df, lokasi):
    email_to_name = buat_mapping_driver(lokasi)
    df = source_df.copy()

    required_cols = {
        "Vehicle Name": "", "Assignee": "", "Weight Percentage": "",
        "Volume Percentage": "", "Total Distance (m)": 0, "Total Visits": "",
        "Total Spent Time (mins)": 0
    }
    for col, default in required_cols.items():
        if col not in df.columns:
            df[col] = default

    def parse_to_numeric(value, is_percentage=False):
        try:
            val_str = str(value).strip()
            if is_percentage:
                val_str = val_str.replace('%', '')
            return float(val_str.replace(',', ''))
        except (ValueError, TypeError):
            return 0.0

    df['Weight Percentage'] = df['Weight Percentage'].apply(lambda x: parse_to_numeric(x, is_percentage=True))
    df['Volume Percentage'] = df['Volume Percentage'].apply(lambda x: parse_to_numeric(x, is_percentage=True))
    df['Total Distance (m)'] = df['Total Distance (m)'].apply(parse_to_numeric)
    df['Total Spent Time (mins)'] = df['Total Spent Time (mins)'].apply(parse_to_numeric)

    agg_rules = {
        'Weight Percentage': 'sum',
        'Volume Percentage': 'sum',
        'Total Distance (m)': 'sum',
        'Total Spent Time (mins)': 'sum'
    }
    df_grouped = df.dropna(subset=['Vehicle Name', 'Assignee'])
    df_agg = df_grouped.groupby(['Vehicle Name', 'Assignee']).agg(agg_rules).reset_index()
    df_others = df[df['Vehicle Name'].isna() | df['Assignee'].isna()]
    df = pd.concat([df_agg, df_others], ignore_index=True)

    def to_h_mm(minutes):
        try:
            minutes = float(str(minutes).replace(",", "").strip())
            hours = int(minutes // 60)
            mins = int(round(minutes % 60))
            return f"'{hours}:{mins:02d}"
        except (ValueError, TypeError):
            return ""

    def format_percentage(value):
        try:
            val_float = float(value)
            return f"{val_float:.1f}%"
        except (ValueError, TypeError):
            return ""

    df['Ship Duration'] = df['Total Spent Time (mins)'].apply(to_h_mm)
    df['Weight Percentage'] = df['Weight Percentage'].apply(format_percentage)
    df['Volume Percentage'] = df['Volume Percentage'].apply(format_percentage)
    df['Total Distance (m)'] = df['Total Distance (m)'].astype(int)

    df["Total Visits"] = ""
    df["Total Delivered"] = ""

    df["Assignee"] = df["Assignee"].str.lower().map(email_to_name).fillna(df["Assignee"])

    existing_drivers = set(df["Assignee"].dropna())
    all_drivers = set(email_to_name.values())
    missing_drivers = sorted(list(all_drivers - existing_drivers))
    if missing_drivers:
        missing_df = pd.DataFrame(missing_drivers, columns=["Assignee"])
        df = pd.concat([df, missing_df], ignore_index=True)

    df.sort_values(by="Assignee", inplace=True, na_position='last')
    final_cols = ["Vehicle Name", "Assignee", "Weight Percentage", "Volume Percentage", "Total Distance (m)", "Total Visits", "Total Delivered", "Ship Duration"]
    sheet_detail = workbook.create_sheet(title="Truck Detail")
    sheet_detail.append(final_cols)

    for col in final_cols:
        if col not in df.columns:
            df[col] = ""
    for r in df[final_cols].to_records(index=False):
        sheet_detail.append(list(r))

    center_align_cols = ["Weight Percentage", "Volume Percentage", "Total Distance (m)", "Total Visits", "Total Delivered", "Ship Duration"]
    for col_idx, col_title in enumerate(final_cols, 1):
        max_length = len(str(col_title))
        for row_idx in range(2, sheet_detail.max_row + 1):
            cell_value = sheet_detail.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        col_letter = get_column_letter(col_idx)
        sheet_detail.column_dimensions[col_letter].width = max_length + 2
        alignment = Alignment(horizontal='center') if col_title in center_align_cols else Alignment(horizontal='left')
        for cell in sheet_detail[col_letter]:
            cell.alignment = alignment

    df['Type'] = df['Assignee'].str.extract(r'(DRY|FRZ)', expand=False, flags=re.IGNORECASE).fillna('OTHER')
    distance_summary = df.groupby('Type')['Total Distance (m)'].sum() / 1000
    sheet_dist = workbook.create_sheet(title="Total Distance Summary")
    sheet_dist["A1"] = "DRY"; sheet_dist["B1"] = "FRZ"
    sheet_dist["A2"] = round(distance_summary.get("DRY", 0), 2)
    sheet_dist["B2"] = round(distance_summary.get("FRZ", 0), 2)

def proses_truck_usage(workbook, source_df):
    master = load_master_data()
    df_master = master["df"]
    upload_df = source_df.copy()
    upload_df.drop_duplicates(subset=['Vehicle Name', 'Assignee'], inplace=True, ignore_index=True)
    type_map = load_type_map().get("type", {}) if load_type_map() else {}
    plat_type_map = dict(zip(df_master["Plat"].astype(str), df_master["Type"]))

    def find_vehicle_tag(vehicle_name):
        if pd.isna(vehicle_name):
            return ""
        for plat, v_type in plat_type_map.items():
            if plat in str(vehicle_name):
                # cek substitusi dari type_map
                return type_map.get(v_type.upper(), v_type.upper())
        return ""

    upload_df["Vehicle Tags"] = upload_df["Vehicle Name"].apply(find_vehicle_tag)
    upload_df["Vehicle Tags"] = upload_df["Vehicle Tags"].str.upper()

    # Pisahkan DRY / FROZEN
    dry_df = upload_df[upload_df["Vehicle Tags"].str.contains("DRY", na=False)]
    frozen_df = upload_df[upload_df["Vehicle Tags"].str.contains("FROZEN", na=False)]

    display_order = constants.get("vehicle_types", [])
    counting_order = ["CDE-LONG", "CDD-LONG", "L300", "CDE", "CDD", "FUSO"]

    def count_types(df_tags):
        counts = {v_type: 0 for v_type in display_order}
        tags_list = df_tags.tolist()
        for v_type in counting_order:
            found_count = 0
            remaining_tags = []
            for tag in tags_list:
                if pd.notna(tag) and v_type in tag:
                    found_count += 1
                else:
                    remaining_tags.append(tag)
            counts[v_type] = found_count
            tags_list = remaining_tags
        return counts

    dry_counts = count_types(dry_df["Vehicle Tags"].astype(str))
    frozen_counts = count_types(frozen_df["Vehicle Tags"].astype(str))

    # Tulis ke sheet
    sheet_usage = workbook.create_sheet(title="Truck Usage")
    sheet_usage["A1"] = "Tipe Kendaraan"
    sheet_usage["B1"] = "Jumlah (DRY)"
    sheet_usage["C1"] = "Jumlah (FROZEN)"
    row = 2
    for v_type in display_order:
        sheet_usage[f"A{row}"] = v_type
        dry_count = dry_counts.get(v_type, 0)
        frozen_count = frozen_counts.get(v_type, 0)
        sheet_usage[f"B{row}"] = dry_count if dry_count != 0 else None
        sheet_usage[f"C{row}"] = frozen_count if frozen_count != 0 else None
        row += 1
    for col_letter in ["A", "B", "C"]:
        sheet_usage.column_dimensions[col_letter].width = 25

def get_adjusted_date_from_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if "Others" in wb.sheetnames:
            ws = wb["Others"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "Created Date" in headers:
                col_idx = headers.index("Created Date") + 1
                for row in ws.iter_rows(min_row=2, max_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                    created_date = row[0]
                    if isinstance(created_date, str):
                        try:
                            created_date = datetime.strptime(created_date.strip(), "%d/%m/%Y")
                        except ValueError:
                            return None
                    elif not isinstance(created_date, datetime):
                        return None

                    # Tambah 1 hari
                    new_date = created_date + timedelta(days=1)
                    if new_date.weekday() == 6:  # Minggu
                        new_date += timedelta(days=1)

                    return new_date.strftime("%d.%m.%Y")
        return None
    except Exception:
        return None

# ==============================================================================
# FUNGSI UTAMA
# ==============================================================================

def main():
    try:
        config = load_config()
        if not config or "lokasi" not in config:
            show_error_message("Dibatalkan", ERROR_MESSAGES["LOCATION_CODE_MISSING"])
            return
        lokasi = config["lokasi"]

        all_data = []
        index = 1
        while True:
            path = pilih_file_excel()
            if not path:
                if index == 1:
                    show_info_message("Dibatalkan", INFO_MESSAGES["CANCELED_BY_USER"])
                    return
                else:
                    break
            skip_rows = 10 if contains_capacity_constraint(path) else 0
            df = pd.read_excel(path, skiprows=skip_rows)
            df.columns = df.columns.str.strip()
            all_data.append(df)
            index += 1
            lanjut = show_ask_message("Konfirmasi", ASK_MESSAGES["ASK_ANOTHER_FILE"])
            if not lanjut:
                break

        if not all_data:
            return

        combined_df = pd.concat(all_data, ignore_index=True)
        required_columns = [
            "Vehicle Name", "Assignee", "Weight Percentage", "Volume Percentage",
            "Total Distance (m)", "Total Visits", "Total Spent Time (mins)"
        ]
        if any(col not in combined_df.columns for col in required_columns):
            show_error_message("Proses Gagal", ERROR_MESSAGES["INVALID_FILE"].format(details="Upload file Export Routing dengan benar!"))
            return

        email_prefixes = combined_df["Assignee"].dropna().str.extract(r'kendaraan\.([^.@]+)', expand=False)
        email_prefixes = email_prefixes.dropna().str.lower().unique()
        if not any(lokasi.lower() in prefix for prefix in email_prefixes):
            show_error_message("Proses Gagal", ERROR_MESSAGES["LOCATION_CODE_MISSING"])
            return

        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)
        proses_truck_detail(output_wb, combined_df, lokasi)
        proses_truck_usage(output_wb, combined_df)

        tanggal_str = get_adjusted_date_from_excel(path) or datetime.now().strftime("%d.%m.%Y")
        location_id = constants.get('location_id', {})
        lokasi_name = next((name for name, code in location_id.items() if code == lokasi), lokasi)
        file_basename = f"{lokasi_name} - Routing Summary - {tanggal_str}"
        save_path = get_save_path(file_basename)
        if save_path:
            output_wb.save(save_path)
            open_file_externally(save_path)
        else:
            show_info_message("Dibatalkan", INFO_MESSAGES["CANCELED_BY_USER"])

    except Exception as e:
        error_message = traceback.format_exc()
        show_error_message("Terjadi Kesalahan", ERROR_MESSAGES["UNKNOWN_ERROR"].format(error_detail=f"{e}\n\n{error_message}"))

if __name__ == "__main__":
    main()
